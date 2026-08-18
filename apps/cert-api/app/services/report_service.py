"""Excel report generation via openpyxl."""

import json
from collections import defaultdict
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.config import REPORTS_DIR
from app.db.postgres import db
from app.services.derivation import compute_status_dimensions
from app.services.wms_service import summarize_stock_rows
from app.utils.logging import log

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FONT_CERT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL_CERT = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
_HEADER_FONT_STOCK = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL_STOCK = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

_STATUS_FILLS: dict[str, PatternFill] = {
    "OK": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "URL_NOT_FOUND": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "INCONSISTENT": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "EXPIRED": PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid"),
}
_EXPIRED_FILL = PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid")

_STATUS_LABELS: dict[str, str] = {
    "OK": "Conforme",
    "INCONSISTENT": "Inconsistente",
    "URL_NOT_FOUND": "Nao Encontrado",
    "API_ERROR": "Erro de API",
    "NO_EXPECTED": "Sem Certificacao",
    "EXPIRED": "Vencido",
}

# Rotulos das dimensoes derivadas — os MESMOS textos que o painel exibe, para o
# Excel poder ser conferido linha a linha contra a tela.
_CERT_STATUS_LABELS: dict[str, str] = {"ATIVO": "Ativo", "ENCERRADO": "Encerrado"}
_SITE_STATUS_LABELS: dict[str, str] = {"CONFORME": "Conforme", "NAO_CONFORME": "Nao conforme"}
_LICENSE_STATUS_LABELS: dict[str, str] = {
    "VALIDO": "Valido",
    "VENCIDO": "Vencido",
    "NAO_APLICAVEL": "Nao aplicavel",
}

_FORMULA_PREFIXES = ("=", "+", "-", "@")

# Rotulos da trava de faturamento. "Nao verificado" existe para o Linx fora do ar:
# um relatorio que dissesse "Nao" nesse caso afirmaria ausencia de trava sem ter
# consultado nada — pior que admitir que nao olhou.
TRAVA_NAO_VERIFICADA = "Nao verificado (Linx indisponivel)"
TRAVA_SEM_MARCA = "Nao verificado (marca sem Linx)"

# O Linx usa 01/01/1900 como sentinela de "campo criado, data nao preenchida" —
# e e a MAIORIA das linhas: 2.210 dos 2.400 valores da propriedade de validade
# da Puket e 999 dos 1.256 da Imaginarium (medido em 2026-08-07). Tratar essas
# linhas como trava faria o relatorio afirmar "Sim" para centenas de itens que
# nao travam nada. Nenhum certificado vivo tem validade anterior a 2000.
_TRAVA_ANO_MINIMO = 2000


def _trava_ativa(valor: str | None) -> str | None:
    """Devolve a data da trava quando ela e real; None quando e sentinela/vazio."""
    texto = (valor or "").strip()
    if not texto:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            if datetime.strptime(texto, fmt).year < _TRAVA_ANO_MINIMO:
                return None
            return texto
        except ValueError:
            continue
    # Texto que nao e data: nao da para afirmar que trava, mas tambem nao da para
    # descartar — devolve como veio para a operacao julgar.
    return texto


def _fetch_stock_map() -> dict[str, dict]:
    """Fetch aggregated stock totals from cert_stock, per SKU.

    Usa `summarize_stock_rows` — a mesma funcao do painel — para que os totais do
    Excel e da tela nao possam divergir (ver o docstring dela para o historico).

    Returns:
        Dict {sku: {stock_cd, stock_ecommerce, stock_total, stock_synced_at, ...}}.
    """
    try:
        with db() as (_conn, _cur):
            _cur.execute("""
                SELECT sku, source, warehouse,
                    COALESCE(SUM(quantity), 0) AS quantity,
                    COALESCE(SUM(available), 0) AS available,
                    MAX(synced_at) AS synced_at
                FROM cert_stock GROUP BY sku, source, warehouse
            """)
            return summarize_stock_rows([dict(r) for r in _cur.fetchall()])
    except Exception as e:
        log.warning(f"Could not fetch stock data: {e}")
        return {}


def _fetch_travas_faturamento(rows: list[dict]) -> dict[str, dict[str, str]]:
    """Descobre, por SKU, se a trava de faturamento esta gravada no Linx.

    A "trava" e a data escrita na propriedade do produto no ERP: VALIDADE DO
    CERTIFICADO (00224 Puket / 00106 Imaginarium) e VENCIMENTO DO LICENCIAMENTO
    (00225 / 00107). Enquanto ela estiver la, o item nao fatura depois do prazo.

    A fonte e o PROPRIO Linx, nao o portal. A primeira versao lia
    `cert_certificates` — a tabela do formulario de cadastro de certificado — e
    por isso reportava "sem trava" para 100% dos produtos: a tabela estava vazia
    (ninguem usou o formulario) enquanto o Linx tinha trava para 489 dos 658
    produtos do painel. As travas foram gravadas por outros caminhos, entre eles
    `sync_prazo_venda_to_linx`, que escreve direto no ERP.

    Falha de conexao NAO derruba o relatorio: os SKUs daquela marca saem como
    "Nao verificado" e o Excel e gerado do mesmo jeito.

    Args:
        rows: linhas de cert_products (usa `sku` e `brand`).

    Returns:
        Dict {sku: {'cert': <rotulo>, 'lic': <rotulo>}}.
    """
    from app.db.sqlserver import _brand_linx, fetch_produto_propriedades

    por_marca: dict[str, list[str]] = defaultdict(list)
    for r in rows:
        sku = str(r.get("sku") or "").strip()
        brand = str(r.get("brand") or "").strip()
        if sku and brand:
            por_marca[brand].append(sku)

    travas: dict[str, dict[str, str]] = {}
    for brand, skus in por_marca.items():
        try:
            cfg = _brand_linx(brand)
        except ValueError:
            for sku in skus:
                travas[sku] = {"cert": TRAVA_SEM_MARCA, "lic": TRAVA_SEM_MARCA}
            continue

        prop_cert = cfg["prop_validade_certificado"]
        prop_lic = cfg["prop_vencimento_licenciamento"]
        try:
            props = fetch_produto_propriedades(brand, [prop_cert, prop_lic], skus)
        except Exception as e:
            log.warning(f"Trava de faturamento indisponivel para '{brand}': {e}")
            for sku in skus:
                travas[sku] = {"cert": TRAVA_NAO_VERIFICADA, "lic": TRAVA_NAO_VERIFICADA}
            continue

        for sku in skus:
            valores = props.get(sku, {})
            # O valor da propriedade E a data da trava — vai junto no rotulo,
            # porque saber ate quando o item fatura vale mais que um "Sim".
            # `_trava_ativa` descarta a sentinela 01/01/1900 do Linx.
            data_cert = _trava_ativa(valores.get(prop_cert))
            data_lic = _trava_ativa(valores.get(prop_lic))
            travas[sku] = {
                "cert": f"Sim ({data_cert})" if data_cert else "Nao",
                "lic": f"Sim ({data_lic})" if data_lic else "Nao",
            }

    log.info(f"Travas de faturamento consultadas no Linx para {len(travas)} SKUs")
    return travas


def _row_get(row: Any, key: str, default: Any = "") -> Any:
    """Return a column from dict-like, RealDictRow, tuple, or object rows."""
    if hasattr(row, "get"):
        return row.get(key, default)
    if isinstance(row, tuple) and hasattr(row, "_fields") and key in row._fields:
        return getattr(row, key)
    return getattr(row, key, default)


def _safe_int(value: Any) -> int:
    """Normalize DB numeric values for Excel cells."""
    if value is None or value == "":
        return 0
    if isinstance(value, Decimal):
        return int(value)
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _safe_text(value: Any) -> str:
    """Normalize nullable values and neutralize formula-like text for Excel cells."""
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        text = value.isoformat()
    else:
        text = str(value)
    stripped = text.lstrip(" \t\r\n")
    if stripped.startswith(_FORMULA_PREFIXES):
        return f"'{text}"
    return text


def _apply_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: list[str],
    header_font: Font,
    header_fill: PatternFill,
) -> int:
    """Write and style the header row.

    Args:
        ws: Active worksheet.
        headers: List of column header labels.
        header_font: Font to apply to header cells.
        header_fill: Fill to apply to header cells.

    Returns:
        Row index of the header row.
    """
    ws.append(headers)
    header_row = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER
    return header_row


# Colunas do relatorio de produtos, na ordem. A lista existe para o cabecalho, a
# largura e a montagem da linha ficarem sempre em sincronia.
_PRODUCT_COLUMNS: tuple[tuple[str, int], ...] = (
    ("SKU", 15),
    ("Nome", 40),
    ("Marca", 18),
    # As tres dimensoes do painel, com os mesmos rotulos da tela.
    ("Status Certificacao", 18),
    ("Status E-commerce", 18),
    ("Motivo (E-commerce)", 38),
    ("Status Licenciamento", 20),
    # Cadastro vindo das abas Imaginarium/Puket.
    ("Tipo Certificacao", 32),          # coluna H
    ("Numero Certificado", 20),         # coluna P
    ("Texto Esperado", 45),             # coluna V
    ("Texto Encontrado", 45),
    ("Pontuacao", 11),
    ("URL", 45),
    # Aba Encerramentos.
    ("Prazo Final Venda", 16),          # coluna G
    ("Situacao da Venda", 26),          # coluna H
    ("Licen. - Prazo", 16),
    # Travas de faturamento gravadas no Linx.
    ("Trava Fat. Certificacao", 26),
    ("Trava Fat. Licenciamento", 26),
    # Estoque.
    ("Estoque CD Disponivel", 20),
    ("Estoque E-commerce", 18),
    ("Total Estoque", 14),
    ("Estoque Atualizado Em", 22),
)

# 1-based, usada para pintar a celula de status de certificacao.
_COL_STATUS_CERT = 4

def generate_products_report(
    rows: list[dict],
    brand: str = "",
    status: str = "",
    license_map: dict | None = None,
) -> Path:
    """Generate an Excel report for cert_products data.

    O relatorio espelha o painel: as colunas Status Certificacao / Status
    E-commerce / Status Licenciamento saem de `compute_status_dimensions`, a
    MESMA funcao que alimenta a tela, em vez de reexibirem o
    `last_validation_status` cru (que so fala do scraping da VTEX e nao do
    veredito de negocio).

    Args:
        rows: List of product dicts from cert_products.
        brand: Optional brand filter label (used only in filename).
        status: Optional status filter label (used only in filename).
        license_map: Mapa de "Licenciamentos Vencidos" (SKU -> status/prazo).
            Sem ele o status de licenciamento sai como "Nao aplicavel".

    Returns:
        Path to the generated .xlsx file.
    """
    now = datetime.now(UTC)
    stock_map = _fetch_stock_map()
    travas = _fetch_travas_faturamento(rows)

    enriched = [{**r, **compute_status_dimensions(r, license_map)} for r in rows]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"

    # Meta rows — contagens pelas dimensoes de negocio, nao pelo status cru.
    ativos = sum(1 for r in enriched if r.get("cert_status") == "ATIVO")
    encerrados = sum(1 for r in enriched if r.get("cert_status") == "ENCERRADO")
    conformes = sum(1 for r in enriched if r.get("site_status") == "CONFORME")
    nao_conformes = sum(1 for r in enriched if r.get("site_status") == "NAO_CONFORME")
    lic_vencidos = sum(1 for r in enriched if r.get("license_status") == "VENCIDO")
    bloqueados = sum(1 for r in enriched if r.get("venda_encerramento") == "BLOQUEADA")

    ws.append(["Relatorio de Produtos - Certificacoes"])
    ws.merge_cells("A1:J1")
    ws["A1"].font = Font(bold=True, size=14, color="059669")
    ws.append([f"Gerado em: {now.strftime('%d/%m/%Y %H:%M')}"])
    ws.append([f"Total: {len(rows)} produtos"])
    ws.append([])
    ws.append(
        [
            f"Certificacao — Ativo: {ativos} | Encerrado: {encerrados}    "
            f"E-commerce — Conforme: {conformes} | Nao conforme: {nao_conformes}    "
            f"Licenciamento vencido: {lic_vencidos}    Venda bloqueada: {bloqueados}"
        ]
    )
    ws.append([_estoque_meta_line(stock_map)])

    headers = [name for name, _ in _PRODUCT_COLUMNS]
    header_row = _apply_header_row(ws, headers, _HEADER_FONT_CERT, _HEADER_FILL_CERT)

    for r in enriched:
        cert_status = r.get("cert_status") or ""
        score = r.get("last_validation_score")
        score_str = f"{score * 100:.0f}%" if score is not None else ""
        sku = r.get("sku", "")
        stock = stock_map.get(sku, {})
        trava = travas.get(sku, {})
        row_data = [
            _safe_text(sku),
            _safe_text(r.get("name", "")),
            _safe_text(r.get("brand", "")),
            _safe_text(_CERT_STATUS_LABELS.get(cert_status, cert_status)),
            _safe_text(_SITE_STATUS_LABELS.get(r.get("site_status") or "", r.get("site_status") or "")),
            _safe_text(r.get("site_status_reason") or ""),
            _safe_text(
                _LICENSE_STATUS_LABELS.get(
                    r.get("license_status") or "", r.get("license_status") or ""
                )
            ),
            _safe_text(r.get("certification_type", "")),
            _safe_text(r.get("numero_certificado", "")),
            _safe_text(r.get("expected_cert_text", "")),
            _safe_text(r.get("actual_cert_text", "")),
            _safe_text(score_str),
            _safe_text(r.get("last_validation_url", "")),
            _safe_text(r.get("sale_deadline", "")),
            _safe_text(r.get("encerramento_status", "")),
            _safe_text(r.get("license_deadline") or ""),
            _safe_text(trava.get("cert", TRAVA_NAO_VERIFICADA)),
            _safe_text(trava.get("lic", TRAVA_NAO_VERIFICADA)),
            stock.get("stock_cd", 0),
            stock.get("stock_ecommerce", 0),
            stock.get("stock_total", 0),
            _safe_text(stock.get("stock_synced_at") or ""),
        ]
        ws.append(row_data)
        row_idx = ws.max_row
        for col_idx in range(1, len(row_data) + 1):
            ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER
        status_cell = ws.cell(row=row_idx, column=_COL_STATUS_CERT)
        if cert_status == "ENCERRADO":
            status_cell.fill = _EXPIRED_FILL
        elif cert_status == "ATIVO":
            status_cell.fill = _STATUS_FILLS["OK"]

    for i, (_, width) in enumerate(_PRODUCT_COLUMNS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    last_col = openpyxl.utils.get_column_letter(len(_PRODUCT_COLUMNS))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"

    filename = f"produtos_certificacoes_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = REPORTS_DIR / filename
    wb.save(str(filepath))
    return filepath


def _estoque_meta_line(stock_map: dict[str, dict]) -> str:
    """Linha de cabecalho que datalha o estoque — o relatorio nao pode esconder
    que os numeros vem de um sync antigo.

    O sync de estoque nao tinha agendamento e ficou 4 meses sem rodar (ultimo em
    23/03/2026, descoberto em 2026-08-07). Quem abrir o Excel precisa ver a data.
    """
    datas = [s.get("stock_synced_at") for s in stock_map.values() if s.get("stock_synced_at")]
    if not datas:
        return "Estoque: sem dados sincronizados"
    return f"Estoque sincronizado em: {max(datas)[:16].replace('T', ' ')}"


def generate_stock_report(rows: list, brand: str = "") -> Path:
    """Generate a detailed stock Excel report.

    Args:
        rows: Raw rows from cert_stock JOIN cert_products query.
        brand: Optional brand filter label.

    Returns:
        Path to the generated .xlsx file.
    """
    now = datetime.now(UTC)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque Detalhado"

    ws.append(["Relatório de Estoque Detalhado - CD Biguaçu + E-commerce"])
    ws.merge_cells("A1:M1")
    ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
    ws.append([f"Data: {now.strftime('%d/%m/%Y %H:%M')}"])
    ws.append([f"Total registros: {len(rows)}"])
    ws.append([])

    headers = [
        "SKU",
        "Nome",
        "Marca",
        "Origem",
        "Localização",
        "Quantidade",
        "Disponível",
        "Reserva",
        "Trânsito",
        "Situação",
        "Status Cert",
        "Prazo Venda",
        "Sincronizado em",
    ]
    header_row = _apply_header_row(ws, headers, _HEADER_FONT_STOCK, _HEADER_FILL_STOCK)

    source_labels = {
        "wms_biguacu": "CD Biguaçu (WMS)",
        "ecommerce_puket": "E-commerce Puket",
        "ecommerce_imaginarium": "E-commerce Imaginarium",
    }
    wms_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    ecom_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    for row_data in rows:
        source_raw = _safe_text(_row_get(row_data, "source", ""))
        row_values = [
            _safe_text(_row_get(row_data, "sku", "")),
            _safe_text(_row_get(row_data, "name", "")),
            _safe_text(_row_get(row_data, "brand", "")),
            source_labels.get(source_raw, source_raw),
            _safe_text(_row_get(row_data, "warehouse", "")).replace("CD ", ""),
            _safe_int(_row_get(row_data, "quantity", 0)),
            _safe_int(_row_get(row_data, "available", 0)),
            _safe_int(_row_get(row_data, "reserved", 0)),
            _safe_int(_row_get(row_data, "in_transit", 0)),
            _safe_text(_row_get(row_data, "situation", "")),
            _safe_text(_row_get(row_data, "last_validation_status", "")),
            _safe_text(_row_get(row_data, "sale_deadline", "")),
            _safe_text(_row_get(row_data, "synced_at", "")),
        ]
        ws.append(row_values)
        row_idx = ws.max_row
        fill = wms_fill if "wms" in source_raw else ecom_fill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER
            cell.fill = fill

    col_widths = [15, 45, 18, 25, 22, 12, 12, 10, 10, 20, 15, 14, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A{header_row}:M{ws.max_row}"

    filename = f"estoque_detalhado_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = REPORTS_DIR / filename
    wb.save(str(filepath))
    return filepath


def generate_validation_report_xlsx(json_filename: str) -> Path:
    """Convert a JSON validation report to Excel format.

    Args:
        json_filename: Filename (basename only) of the .json report in REPORTS_DIR.

    Returns:
        Path to the generated .xlsx file.
    """
    json_path = REPORTS_DIR / json_filename
    report_data = json.loads(json_path.read_text())
    products = report_data.get("products", report_data.get("results", []))
    summary = report_data.get("summary", {})

    stock_map = _fetch_stock_map()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validação"

    ws.append(["Relatório de Validação de Certificações"])
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=14, color="059669")
    ws.append([f"Data: {report_data.get('date', '')}"])
    ws.append(
        [
            f"Total: {summary.get('total', len(products))} | OK: {summary.get('ok', 0)} "
            f"| Ausente: {summary.get('missing', 0)} | Inconsistente: {summary.get('inconsistent', 0)} "
            f"| Não Encontrado: {summary.get('not_found', 0)}"
        ]
    )
    ws.append([])

    headers = [
        "SKU",
        "Nome",
        "Marca",
        "Status",
        "Pontuacao",
        "Tipo Certificacao",
        "Texto Esperado",
        "Texto Encontrado",
        "URL",
        "Erro",
        "Estoque CD Disponivel",
        "Estoque E-commerce",
        "Total Estoque",
    ]
    header_row = _apply_header_row(ws, headers, _HEADER_FONT_CERT, _HEADER_FILL_CERT)

    for p in products:
        status_raw = p.get("status", "")
        status_label = _STATUS_LABELS.get(status_raw, status_raw)
        score = p.get("score")
        score_str = f"{score * 100:.0f}%" if score is not None else ""
        p_sku = p.get("sku", "")
        stock = stock_map.get(p_sku, {})
        row = [
            _safe_text(p_sku),
            _safe_text(p.get("name", "")),
            _safe_text(p.get("brand", "")),
            _safe_text(status_label),
            _safe_text(score_str),
            _safe_text(p.get("certification_type", "")),
            _safe_text(p.get("expected_cert_text", "")),
            _safe_text(p.get("actual_cert_text", "")),
            _safe_text(p.get("url", "")),
            _safe_text(p.get("error", "")),
            stock.get("stock_cd", 0),
            stock.get("stock_ecommerce", 0),
            stock.get("stock_total", 0),
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(row) + 1):
            ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER
        status_cell = ws.cell(row=row_idx, column=4)
        if status_raw in _STATUS_FILLS:
            status_cell.fill = _STATUS_FILLS[status_raw]

    col_widths = [15, 40, 18, 18, 12, 32, 40, 40, 50, 40, 12, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A{header_row}:M{ws.max_row}"

    excel_filename = json_filename.replace(".json", ".xlsx")
    excel_path = REPORTS_DIR / excel_filename
    wb.save(str(excel_path))
    return excel_path
