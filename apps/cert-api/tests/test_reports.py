"""Report route and XLSX generation tests."""

from datetime import UTC, datetime

import openpyxl
import pytest

from app.services.report_service import generate_products_report, generate_stock_report


def _patch_reports_db(mocker, rows=None):
    cursor = mocker.MagicMock()
    cursor.fetchall.return_value = rows or []
    conn = mocker.MagicMock()
    ctx = mocker.MagicMock()
    ctx.__enter__ = mocker.MagicMock(return_value=(conn, cursor))
    ctx.__exit__ = mocker.MagicMock(return_value=False)
    mocker.patch("app.routes.reports.db", return_value=ctx)
    return cursor


@pytest.mark.asyncio
async def test_export_stock_normalizes_brand_and_keeps_wms_join(test_client, api_key_headers, mocker, tmp_path):
    """Stock export should filter by product brand alias, not raw cert_stock.brand."""
    mocker.patch("app.routes.reports.DATABASE_URL", "postgres://test")
    cursor = _patch_reports_db(mocker, rows=[{"sku": "PI4257Y", "source": "wms_biguacu"}])
    output = tmp_path / "estoque.xlsx"
    output.write_bytes(b"PK\x03\x04xlsx")
    mocker.patch("app.routes.reports.generate_stock_report", return_value=output)

    resp = await test_client.post(
        "/api/reports/export-stock?brand=puket_escolares",
        headers=api_key_headers,
    )

    assert resp.status_code == 200
    sql, params = cursor.execute.call_args.args
    assert "COALESCE(cp.brand, cs.brand" in sql
    assert params == ["puket escolares"]


@pytest.mark.asyncio
async def test_export_stock_permission_error_is_actionable(test_client, api_key_headers, mocker):
    """Permission errors on cert-reports should return a clear operational message."""
    mocker.patch("app.routes.reports.DATABASE_URL", "postgres://test")
    _patch_reports_db(mocker)
    mocker.patch("app.routes.reports.generate_stock_report", side_effect=PermissionError())

    resp = await test_client.post("/api/reports/export-stock", headers=api_key_headers)

    assert resp.status_code == 500
    assert "cert-reports" in resp.json()["detail"]


@pytest.mark.asyncio
async def test_reports_list_includes_format(test_client, api_key_headers, mocker, tmp_path):
    """The frontend uses format to hide JSON-only actions for generated XLSX files."""
    mocker.patch("app.routes.reports.REPORTS_DIR", tmp_path)
    (tmp_path / "validation.json").write_text("{}", encoding="utf-8")
    (tmp_path / "estoque.xlsx").write_bytes(b"PK\x03\x04xlsx")

    resp = await test_client.get("/api/reports", headers=api_key_headers)

    assert resp.status_code == 200
    formats = {item["filename"]: item["format"] for item in resp.json()}
    assert formats["validation.json"] == "json"
    assert formats["estoque.xlsx"] == "xlsx"


@pytest.mark.asyncio
async def test_report_data_rejects_xlsx(test_client, api_key_headers, mocker, tmp_path):
    """XLSX files are binary and should not be parsed through the JSON detail endpoint."""
    mocker.patch("app.routes.reports.REPORTS_DIR", tmp_path)
    (tmp_path / "estoque.xlsx").write_bytes(b"PK\x03\x04xlsx")

    resp = await test_client.get("/api/reports/estoque.xlsx/data", headers=api_key_headers)

    assert resp.status_code == 400
    assert "JSON" in resp.json()["detail"]


def test_generate_stock_report_writes_synced_at_column(mocker, tmp_path):
    """Generated stock XLSX should include the sync timestamp for WMS auditability."""
    mocker.patch("app.services.report_service.REPORTS_DIR", tmp_path)
    output = generate_stock_report(
        [
            {
                "sku": "PI4257Y",
                "name": "Produto",
                "brand": "Puket",
                "source": "wms_biguacu",
                "warehouse": "CD Picking",
                "quantity": 10,
                "available": 8,
                "reserved": 1,
                "in_transit": 2,
                "situation": "LIBERADO",
                "last_validation_status": "OK",
                "sale_deadline": "2026-12-31",
                "synced_at": datetime(2026, 6, 19, 12, 0, tzinfo=UTC),
            }
        ]
    )

    wb = openpyxl.load_workbook(output)
    ws = wb["Estoque Detalhado"]
    headers = [cell.value for cell in ws[5]]
    assert "Sincronizado em" in headers
    assert ws["M6"].value.startswith("2026-06-19T12:00:00")


def _patch_products_report_io(mocker, tmp_path, stock=None, travas=None):
    """Isola a geracao do XLSX de produtos das consultas ao banco."""
    mocker.patch("app.services.report_service.REPORTS_DIR", tmp_path)
    mocker.patch("app.services.report_service._fetch_stock_map", return_value=stock or {})
    mocker.patch("app.services.report_service._fetch_travas_faturamento", return_value=travas or {})


def _header_index(ws, label: str) -> int:
    """Indice 1-based da coluna cujo cabecalho (linha 7) e `label`."""
    headers = [c.value for c in ws[7]]
    return headers.index(label) + 1


def test_generated_xlsx_neutralizes_formula_like_text(mocker, tmp_path):
    """User-controlled text exported to XLSX must not execute as formulas."""
    _patch_products_report_io(mocker, tmp_path)

    output = generate_products_report(
        [
            {
                "sku": "=2+2",
                "name": "+Produto",
                "brand": "@Marca",
                "last_validation_status": "OK",
                "certification_type": "-Tipo",
                "numero_certificado": "=Cert",
                "expected_cert_text": "=Esperado",
                "actual_cert_text": "+Encontrado",
                "last_validation_url": "@https://example.invalid",
                "sale_deadline": "-2026-12-31",
                "encerramento_status": "@Comerciacao Permitida",
            }
        ]
    )

    wb = openpyxl.load_workbook(output, data_only=False)
    ws = wb["Produtos"]

    def cell(label: str):
        return ws.cell(row=8, column=_header_index(ws, label)).value

    assert cell("SKU") == "'=2+2"
    assert cell("Nome") == "'+Produto"
    assert cell("Marca") == "'@Marca"
    assert cell("Tipo Certificacao") == "'-Tipo"
    assert cell("Numero Certificado") == "'=Cert"
    assert cell("Texto Esperado") == "'=Esperado"
    assert cell("Texto Encontrado") == "'+Encontrado"
    assert cell("URL") == "'@https://example.invalid"
    assert cell("Prazo Final Venda") == "'-2026-12-31"
    assert cell("Situacao da Venda") == "'@Comerciacao Permitida"


def test_products_report_mirrors_panel_status_columns(mocker, tmp_path):
    """O Excel tem de trazer os MESMOS tres status do painel, nao o status cru."""
    _patch_products_report_io(
        mocker,
        tmp_path,
        stock={"PI7560Y": {"stock_cd": 10, "stock_ecommerce": 5, "stock_total": 15,
                           "stock_synced_at": "2026-08-07T09:00:00"}},
    )

    output = generate_products_report(
        [
            {
                "sku": "PI7560Y",
                "name": "CANETA PANDA AMIGOS",
                "brand": "Imaginarium",
                "sheet_status": "27/10/25 - Item excluído e incluído novamente com o novo nome.",
                "encerramento_status": "Comerciação Permitida",
                "last_validation_status": "OK",
            }
        ],
        license_map={"PI7560Y": {"status": "VENCIDO", "valid_until": "2026-01-31"}},
    )

    wb = openpyxl.load_workbook(output)
    ws = wb["Produtos"]

    def cell(label: str):
        return ws.cell(row=8, column=_header_index(ws, label)).value

    assert cell("Status Certificacao") == "Ativo"
    assert cell("Status E-commerce") == "Conforme"
    assert cell("Status Licenciamento") == "Vencido"
    assert cell("Licen. - Prazo") == "2026-01-31"
    assert cell("Estoque CD") == 10
    assert cell("Total Estoque") == 15
    assert cell("Estoque Atualizado Em") == "2026-08-07T09:00:00"


def test_products_report_has_no_vencido_column_and_reports_travas(mocker, tmp_path):
    """A coluna 'Vencido' saiu; entraram as duas travas de faturamento."""
    _patch_products_report_io(
        mocker,
        tmp_path,
        travas={"PI7223Y": {"cert": "Sim (24/07/2026)", "lic": "Nao - sem data cadastrada"}},
    )

    output = generate_products_report([{"sku": "PI7223Y", "name": "CAIXA DE SOM", "brand": "Imaginarium"}])

    wb = openpyxl.load_workbook(output)
    ws = wb["Produtos"]
    headers = [c.value for c in ws[7]]

    assert "Vencido" not in headers
    assert ws.cell(row=8, column=_header_index(ws, "Trava Fat. Certificacao")).value == "Sim (24/07/2026)"
    assert (
        ws.cell(row=8, column=_header_index(ws, "Trava Fat. Licenciamento")).value
        == "Nao - sem data cadastrada"
    )


def test_products_report_marks_missing_certificate_registration(mocker, tmp_path):
    """SKU que o Linx nao respondeu nao pode parecer 'sem trava'."""
    _patch_products_report_io(mocker, tmp_path)

    output = generate_products_report([{"sku": "PI9999Y", "name": "X", "brand": "Puket"}])

    wb = openpyxl.load_workbook(output)
    ws = wb["Produtos"]
    assert (
        ws.cell(row=8, column=_header_index(ws, "Trava Fat. Certificacao")).value
        == "Nao verificado (Linx indisponivel)"
    )


class TestTravaFaturamento:
    """A trava vem do Linx (PROP_PRODUTOS), nunca de cert_certificates.

    A primeira versao lia a tabela do formulario do portal e reportava "sem
    trava" nas 658 linhas, enquanto o Linx tinha trava para 489 produtos.
    """

    def _rows(self):
        return [
            {"sku": "PI7223Y", "brand": "Imaginarium"},
            {"sku": "100400496", "brand": "Puket"},
        ]

    def test_le_a_propriedade_do_linx_e_mostra_a_data(self, mocker):
        from app.services import report_service

        mocker.patch(
            "app.db.sqlserver.fetch_produto_propriedades",
            side_effect=lambda brand, props, skus: (
                {"PI7223Y": {"00106": "24/07/2026", "00107": "31/12/2026"}}
                if brand == "Imaginarium"
                else {"100400496": {"00224": "11/08/2027"}}
            ),
        )
        travas = report_service._fetch_travas_faturamento(self._rows())

        assert travas["PI7223Y"] == {"cert": "Sim (24/07/2026)", "lic": "Sim (31/12/2026)"}
        # Produto com validade mas sem licenciamento gravado.
        assert travas["100400496"] == {"cert": "Sim (11/08/2027)", "lic": "Nao"}

    def test_produto_sem_propriedade_sai_como_nao(self, mocker):
        from app.services import report_service

        mocker.patch("app.db.sqlserver.fetch_produto_propriedades", return_value={})
        travas = report_service._fetch_travas_faturamento(self._rows())
        assert travas["PI7223Y"] == {"cert": "Nao", "lic": "Nao"}

    def test_linx_fora_do_ar_nao_derruba_o_relatorio(self, mocker):
        """Dizer "Nao" sem ter consultado seria afirmar ausencia de trava."""
        from app.services import report_service

        mocker.patch(
            "app.db.sqlserver.fetch_produto_propriedades",
            side_effect=OSError("db02 unreachable"),
        )
        travas = report_service._fetch_travas_faturamento(self._rows())
        assert travas["PI7223Y"]["cert"] == report_service.TRAVA_NAO_VERIFICADA
        assert travas["100400496"]["lic"] == report_service.TRAVA_NAO_VERIFICADA

    def test_marca_sem_linx_e_sinalizada(self, mocker):
        from app.services import report_service

        mocker.patch("app.db.sqlserver.fetch_produto_propriedades", return_value={})
        travas = report_service._fetch_travas_faturamento([{"sku": "X1", "brand": "Kayuan"}])
        assert travas["X1"]["cert"] == report_service.TRAVA_SEM_MARCA

    def test_consulta_uma_vez_por_marca_e_nao_por_sku(self, mocker):
        """658 produtos nao podem virar 658 idas ao SQL Server."""
        from app.services import report_service

        spy = mocker.patch("app.db.sqlserver.fetch_produto_propriedades", return_value={})
        report_service._fetch_travas_faturamento(
            [{"sku": f"PI{i}Y", "brand": "Imaginarium"} for i in range(300)]
        )
        assert spy.call_count == 1


class TestSentinelaDoLinx:
    """01/01/1900 e "campo criado sem data", nao trava — e e a maioria das linhas."""

    def test_sentinela_1900_nao_e_trava(self):
        from app.services.report_service import _trava_ativa

        assert _trava_ativa("01/01/1900") is None
        assert _trava_ativa("1900-01-01") is None

    def test_data_real_e_trava(self):
        from app.services.report_service import _trava_ativa

        assert _trava_ativa("24/07/2026") == "24/07/2026"
        assert _trava_ativa(" 11/08/2027 ") == "11/08/2027"

    def test_vazio_nao_e_trava(self):
        from app.services.report_service import _trava_ativa

        assert _trava_ativa("") is None
        assert _trava_ativa(None) is None

    def test_texto_nao_data_volta_como_veio(self):
        """Nao da para afirmar nem descartar — a operacao julga."""
        from app.services.report_service import _trava_ativa

        assert _trava_ativa("indeterminado") == "indeterminado"

    def test_sentinela_no_lookup_completo(self, mocker):
        from app.services import report_service

        mocker.patch(
            "app.db.sqlserver.fetch_produto_propriedades",
            return_value={"PI1Y": {"00106": "01/01/1900", "00107": "31/12/2027"}},
        )
        travas = report_service._fetch_travas_faturamento([{"sku": "PI1Y", "brand": "Imaginarium"}])
        assert travas["PI1Y"] == {"cert": "Nao", "lic": "Sim (31/12/2027)"}
