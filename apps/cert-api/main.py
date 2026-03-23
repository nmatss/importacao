"""
Cert-API: Microservice for product certification validation.
Validates certification info on VTEX e-commerce sites against Google Sheets data.
"""

import os
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), '..', '..', '.env'))
import re
import hmac
import uuid
import json
import time
import threading
import logging
import difflib
from html import unescape
from datetime import datetime, timezone
from pathlib import Path
from contextlib import contextmanager

import requests
import psycopg2
import psycopg2.extras
from psycopg2 import pool as pg_pool
import gspread
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google.oauth2.service_account import Credentials
from fastapi import FastAPI, HTTPException, Query, Security, Depends, Request
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import APIKeyHeader
from pydantic import BaseModel
from slowapi import Limiter
from slowapi.util import get_remote_address
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

log = logging.getLogger("cert-api")
logging.basicConfig(level=logging.INFO)

# ---------------------------------------------------------------------------
# API Key auth
# ---------------------------------------------------------------------------

API_KEY = os.environ.get("CERT_API_KEY", "")
api_key_header = APIKeyHeader(name="X-API-Key", auto_error=False)


async def verify_api_key(request: Request, api_key: str = Security(api_key_header)):
    """Verify API key from X-API-Key header. Skips auth if CERT_API_KEY is not set."""
    if request.url.path == "/api/health":
        return
    if API_KEY and not hmac.compare_digest(api_key or "", API_KEY):
        raise HTTPException(status_code=403, detail="Invalid API key")

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = FastAPI(
    title="Cert-API",
    version="2.0.0",
    dependencies=[Depends(verify_api_key)],
)

limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter

_cors_origins_env = os.environ.get("CORS_ORIGINS", "")
_cors_origins = (
    [o.strip() for o in _cors_origins_env.split(",") if o.strip()]
    if _cors_origins_env
    else ["http://localhost:5173", "http://localhost:8080"]
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)

DATABASE_URL = os.environ.get("DATABASE_URL", "")
REPORTS_DIR = Path(os.environ.get("REPORTS_DIR", Path(__file__).parent / "reports"))
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# Google Sheets config
SHEETS_CLIENT_EMAIL = os.environ.get("GOOGLE_SHEETS_CLIENT_EMAIL", "") or os.environ.get("GOOGLE_DRIVE_CLIENT_EMAIL", "")
SHEETS_PRIVATE_KEY = (os.environ.get("GOOGLE_SHEETS_PRIVATE_KEY", "") or os.environ.get("GOOGLE_DRIVE_PRIVATE_KEY", "")).replace("\\n", "\n")
SHEETS_SPREADSHEET_ID = os.environ.get(
    "GOOGLE_SHEETS_SPREADSHEET_ID",
    "1qcgcj9814UFikhurgvsTTcUxvPF2r3w_QY_EurvBtSE",
)

# VTEX config — brand → store domain mapping
# Note: "puket escolares" must come before "puket" so exact match wins over substring
VTEX_STORES: dict[str, dict] = {
    "puket escolares": {
        "domain": os.environ.get("VTEX_PUKET_DOMAIN", "www.puket.com.br"),
        "site_url": "https://www.puket.com.br",
    },
    "puket": {
        "domain": os.environ.get("VTEX_PUKET_DOMAIN", "www.puket.com.br"),
        "site_url": "https://www.puket.com.br",
    },
    "imaginarium": {
        "domain": os.environ.get("VTEX_IMAGINARIUM_DOMAIN", "loja.imaginarium.com.br"),
        "site_url": "https://loja.imaginarium.com.br",
    },
}

VTEX_REQUEST_DELAY = float(os.environ.get("VTEX_REQUEST_DELAY", "1.5"))

# Certification keywords to search for in product pages
CERT_KEYWORDS = [
    "inmetro", "certificação", "certificacao", "registro",
    "portaria", "conformidade", "selo", "norma",
    "nbr", "abnt", "anvisa", "certificado", "homologação",
    "homologacao", "regulamento", "oc ", "ocp ",
]

# In-memory store for running validations
_running_validations: dict[str, dict] = {}

def _cleanup_old_validations(max_age_seconds: int = 3600, stuck_timeout: int = 7200):
    """Remove completed/error entries older than max_age_seconds, and stuck 'running' entries older than stuck_timeout."""
    now = time.time()
    to_remove = [
        rid for rid, state in _running_validations.items()
        if (state.get("status") in ("completed", "error")
            and now - state.get("_finished_at", now) > max_age_seconds)
        or (state.get("status") == "running"
            and now - state.get("_started_at", now) > stuck_timeout)
    ]
    for rid in to_remove:
        del _running_validations[rid]

# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

_pool: pg_pool.ThreadedConnectionPool | None = None
_pool_lock = threading.Lock()


def _get_pool() -> pg_pool.ThreadedConnectionPool:
    global _pool
    if _pool is None or _pool.closed:
        with _pool_lock:
            if _pool is None or _pool.closed:
                _pool = pg_pool.ThreadedConnectionPool(1, 10, DATABASE_URL)
    return _pool


def get_conn():
    return _get_pool().getconn()


@contextmanager
def db():
    conn = get_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            yield conn, cur
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        _get_pool().putconn(conn)


def _add_column_if_not_exists(col: str, coltype: str):
    c = None
    try:
        c = get_conn()
        with c.cursor() as cur:
            cur.execute(f"ALTER TABLE cert_products ADD COLUMN IF NOT EXISTS {col} {coltype}")
        c.commit()
    except Exception:
        if c:
            try:
                c.rollback()
            except Exception:
                pass
    finally:
        if c:
            try:
                _get_pool().putconn(c)
            except Exception:
                pass


def ensure_tables():
    """Create cert tables if they don't exist."""
    with db() as (conn, cur):
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cert_products (
                sku TEXT PRIMARY KEY,
                name TEXT NOT NULL DEFAULT '',
                brand TEXT NOT NULL DEFAULT '',
                certification_type TEXT DEFAULT '',
                sheet_status TEXT DEFAULT '',
                expected_cert_text TEXT DEFAULT '',
                ecommerce_description TEXT DEFAULT '',
                actual_cert_text TEXT DEFAULT '',
                last_validation_status TEXT,
                last_validation_score DOUBLE PRECISION,
                last_validation_url TEXT,
                last_validation_date TIMESTAMPTZ,
                last_validation_error TEXT,
                created_at TIMESTAMPTZ DEFAULT NOW(),
                updated_at TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cert_schedules (
                id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                name TEXT NOT NULL,
                brand_filter TEXT,
                cron_expression TEXT NOT NULL,
                enabled BOOLEAN DEFAULT TRUE,
                last_run TIMESTAMPTZ,
                next_run TIMESTAMPTZ,
                created_at TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cert_schedule_history (
                id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                schedule_id UUID REFERENCES cert_schedules(id) ON DELETE CASCADE,
                run_date TIMESTAMPTZ DEFAULT NOW(),
                status TEXT DEFAULT 'completed',
                summary JSONB,
                report_file TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS cert_validation_runs (
                id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                status TEXT DEFAULT 'pending',
                brand_filter TEXT,
                total INTEGER DEFAULT 0,
                processed INTEGER DEFAULT 0,
                ok INTEGER DEFAULT 0,
                missing INTEGER DEFAULT 0,
                inconsistent INTEGER DEFAULT 0,
                not_found INTEGER DEFAULT 0,
                started_at TIMESTAMPTZ DEFAULT NOW(),
                finished_at TIMESTAMPTZ,
                report_file TEXT
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS cert_stock (
                id SERIAL PRIMARY KEY,
                sku TEXT NOT NULL,
                brand TEXT,
                source TEXT NOT NULL,
                warehouse TEXT,
                quantity INTEGER DEFAULT 0,
                available INTEGER DEFAULT 0,
                reserved INTEGER DEFAULT 0,
                in_transit INTEGER DEFAULT 0,
                situation TEXT,
                storage_area TEXT,
                synced_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(sku, source, warehouse)
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS cert_stock_sku_idx ON cert_stock(sku)")

    # Migration: add columns for existing deployments
    for col, coltype in [
        ("certification_type", "TEXT DEFAULT ''"),
        ("sheet_status", "TEXT DEFAULT ''"),
        ("expected_cert_text", "TEXT DEFAULT ''"),
        ("ecommerce_description", "TEXT DEFAULT ''"),
        ("actual_cert_text", "TEXT DEFAULT ''"),
        ("last_validation_error", "TEXT"),
        ("sale_deadline", "TEXT"),
        ("sale_deadline_date", "DATE"),
        ("is_expired", "BOOLEAN DEFAULT FALSE"),
    ]:
        _add_column_if_not_exists(col, coltype)


# ---------------------------------------------------------------------------
# Google Sheets helpers
# ---------------------------------------------------------------------------

def _get_sheets_client() -> gspread.Client | None:
    if not SHEETS_CLIENT_EMAIL or not SHEETS_PRIVATE_KEY:
        log.warning("Google Sheets credentials not configured")
        return None
    try:
        creds = Credentials.from_service_account_info(
            {
                "type": "service_account",
                "client_email": SHEETS_CLIENT_EMAIL,
                "private_key": SHEETS_PRIVATE_KEY,
                "token_uri": "https://oauth2.googleapis.com/token",
            },
            scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
        )
        return gspread.authorize(creds)
    except Exception as e:
        log.error(f"Failed to create Sheets client: {e}")
        return None


def _find_col_by_header(headers: list[str], *candidates: str) -> int | None:
    """Find column index by matching header text (case-insensitive, accent-insensitive)."""
    for i, h in enumerate(headers):
        h_lower = h.lower().strip()
        for c in candidates:
            if c.lower() in h_lower:
                return i
    return None


def _read_products_from_sheets() -> list[dict]:
    client = _get_sheets_client()
    if not client:
        return []
    try:
        spreadsheet = client.open_by_key(SHEETS_SPREADSHEET_ID)
    except Exception as e:
        log.error(f"Failed to open spreadsheet: {e}")
        return []

    products: list[dict] = []
    worksheet_configs = [
        {
            "name": "Imaginarium",
            "sku_col": 2, "name_col": 5, "brand_col": 0,
            "brand_default": "Imaginarium",
            "cert_type_col": 7, "status_col": 9,
        },
        {
            "name": "Puket",
            "sku_col": 2, "name_col": 5, "brand_col": 0,
            "brand_default": "Puket",
            "cert_type_col": 7, "status_col": 9,
        },
        {
            "name": "Puket escolares",
            "sku_col": 0, "name_col": 1, "brand_col": None,
            "brand_default": "Puket Escolares",
            "cert_type_col": 2, "status_col": 6,
        },
    ]

    for cfg in worksheet_configs:
        try:
            ws = spreadsheet.worksheet(cfg["name"])
            rows = ws.get_all_values()
        except Exception as e:
            log.warning(f"Could not read worksheet '{cfg['name']}': {e}")
            continue

        if not rows:
            continue

        # Detect "Descrição E-commerce" column dynamically from header row
        headers = rows[0]
        desc_ecommerce_col = _find_col_by_header(
            headers,
            "descrição e-commerce", "descricao e-commerce",
            "descrição ecommerce", "descricao ecommerce",
            "desc e-commerce", "desc ecommerce",
        )
        if desc_ecommerce_col is not None:
            log.info(f"Worksheet '{cfg['name']}': found 'Descrição E-commerce' at column {desc_ecommerce_col}")

        for row in rows[1:]:
            sku_col = cfg["sku_col"]
            if sku_col >= len(row):
                continue
            raw_sku = str(row[sku_col]).strip()
            if not raw_sku:
                continue

            name = str(row[cfg["name_col"]]).strip() if cfg["name_col"] < len(row) else ""
            if cfg["brand_col"] is not None and cfg["brand_col"] < len(row):
                brand = str(row[cfg["brand_col"]]).strip() or cfg["brand_default"]
            else:
                brand = cfg["brand_default"]
            cert_type = str(row[cfg["cert_type_col"]]).strip() if cfg["cert_type_col"] < len(row) else ""
            sheet_status = str(row[cfg["status_col"]]).strip() if cfg["status_col"] < len(row) else ""

            # Read "Descrição E-commerce" — the exact text expected on the site
            desc_ecommerce = ""
            if desc_ecommerce_col is not None and desc_ecommerce_col < len(row):
                desc_ecommerce = str(row[desc_ecommerce_col]).strip()

            # Handle cells with multiple SKUs separated by newlines
            sku_parts = re.split(r'[\r\n]+', raw_sku)
            for sku in sku_parts:
                sku = sku.strip()
                if not sku:
                    continue
                products.append({
                    "sku": sku,
                    "name": name,
                    "brand": brand,
                    "certification_type": cert_type,
                    "sheet_status": sheet_status,
                    "ecommerce_description": desc_ecommerce,
                })

    log.info(f"Read {len(products)} products from Google Sheets")

    # Read "Encerramentos" tab for expired certifications
    try:
        ws_enc = spreadsheet.worksheet("Encerramentos")
        enc_rows = ws_enc.get_all_values()
        if enc_rows:
            enc_headers = enc_rows[0]
            prazo_col = _find_col_by_header(enc_headers, "prazo final venda", "prazo final", "prazo venda")
            sku_col_enc = _find_col_by_header(enc_headers, "sku", "código", "codigo", "ref")
            name_col_enc = _find_col_by_header(enc_headers, "nome", "produto", "descrição", "descricao")
            brand_col_enc = _find_col_by_header(enc_headers, "marca", "brand")

            if prazo_col is not None and sku_col_enc is not None:
                log.info(f"Worksheet 'Encerramentos': found 'Prazo Final Venda' at column {prazo_col}, SKU at column {sku_col_enc}")
                today = datetime.now().date()
                expired_count = 0
                for row in enc_rows[1:]:
                    if sku_col_enc >= len(row):
                        continue
                    raw_sku = str(row[sku_col_enc]).strip()
                    if not raw_sku:
                        continue

                    prazo_str = str(row[prazo_col]).strip() if prazo_col < len(row) else ""
                    if not prazo_str:
                        continue

                    # Parse date — try common formats
                    prazo_date = None
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
                        try:
                            prazo_date = datetime.strptime(prazo_str, fmt).date()
                            break
                        except ValueError:
                            continue

                    if prazo_date is None:
                        continue

                    is_expired = prazo_date < today
                    name = str(row[name_col_enc]).strip() if name_col_enc is not None and name_col_enc < len(row) else ""
                    brand = str(row[brand_col_enc]).strip() if brand_col_enc is not None and brand_col_enc < len(row) else ""

                    sku_parts = re.split(r'[\r\n]+', raw_sku)
                    for sku in sku_parts:
                        sku = sku.strip()
                        if not sku:
                            continue
                        products.append({
                            "sku": sku,
                            "name": name,
                            "brand": brand,
                            "certification_type": f"ENCERRAMENTO - Prazo: {prazo_str}",
                            "sheet_status": "EXPIRED" if is_expired else "EXPIRING",
                            "ecommerce_description": "",
                            "sale_deadline": prazo_str,
                            "sale_deadline_date": prazo_date.isoformat(),
                            "is_expired": is_expired,
                        })
                        if is_expired:
                            expired_count += 1

                log.info(f"Encerramentos: found {expired_count} expired products out of {len(enc_rows) - 1} rows")
            else:
                log.warning(f"Worksheet 'Encerramentos': could not find required columns (prazo={prazo_col}, sku={sku_col_enc})")
    except gspread.exceptions.WorksheetNotFound:
        log.info("Worksheet 'Encerramentos' not found, skipping expiration check")
    except Exception as e:
        log.warning(f"Error reading 'Encerramentos' worksheet: {e}")

    return products


def sync_sheets_to_db() -> dict:
    products = _read_products_from_sheets()
    if not products:
        return {"synced": 0, "error": "No products found or Sheets not configured"}
    if not DATABASE_URL:
        return {"synced": 0, "error": "Database not configured"}

    try:
        with db() as (conn, cur):
            for p in products:
                ecommerce_desc = p.get("ecommerce_description", "")
                # Use ecommerce_description as expected_cert_text when available,
                # otherwise fall back to certification_type
                expected = ecommerce_desc if ecommerce_desc else p["certification_type"]
                sale_deadline = p.get("sale_deadline")
                sale_deadline_date = p.get("sale_deadline_date")
                is_expired = p.get("is_expired", False)
                cur.execute(
                    """
                    INSERT INTO cert_products (sku, name, brand, certification_type, expected_cert_text,
                                               ecommerce_description, sheet_status, sale_deadline,
                                               sale_deadline_date, is_expired, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                    ON CONFLICT (sku) DO UPDATE SET
                        name = COALESCE(NULLIF(EXCLUDED.name, ''), cert_products.name),
                        brand = COALESCE(NULLIF(EXCLUDED.brand, ''), cert_products.brand),
                        certification_type = CASE WHEN EXCLUDED.sale_deadline IS NOT NULL
                            THEN EXCLUDED.certification_type
                            ELSE COALESCE(NULLIF(EXCLUDED.certification_type, ''), cert_products.certification_type) END,
                        expected_cert_text = COALESCE(NULLIF(EXCLUDED.expected_cert_text, ''), cert_products.expected_cert_text),
                        ecommerce_description = COALESCE(NULLIF(EXCLUDED.ecommerce_description, ''), cert_products.ecommerce_description),
                        sheet_status = CASE WHEN EXCLUDED.sale_deadline IS NOT NULL
                            THEN EXCLUDED.sheet_status
                            ELSE COALESCE(NULLIF(EXCLUDED.sheet_status, ''), cert_products.sheet_status) END,
                        sale_deadline = COALESCE(EXCLUDED.sale_deadline, cert_products.sale_deadline),
                        sale_deadline_date = COALESCE(EXCLUDED.sale_deadline_date, cert_products.sale_deadline_date),
                        is_expired = EXCLUDED.is_expired OR cert_products.is_expired,
                        updated_at = NOW()
                    """,
                    [p["sku"], p["name"], p["brand"], p["certification_type"],
                     expected, ecommerce_desc, p["sheet_status"],
                     sale_deadline, sale_deadline_date, is_expired],
                )
        return {"synced": len(products), "total_rows": len(products)}
    except Exception as e:
        log.error(f"Failed to sync sheets to DB: {e}")
        return {"synced": 0, "error": str(e)}


# ---------------------------------------------------------------------------
# VTEX integration
# ---------------------------------------------------------------------------

def _get_vtex_config(brand: str) -> dict | None:
    brand_lower = brand.lower().strip()
    for key, config in VTEX_STORES.items():
        if key in brand_lower or brand_lower in key:
            return config
    return None


def _strip_html(html_text: str, keep_newlines: bool = False) -> str:
    if not html_text:
        return ""
    if keep_newlines:
        # Replace <br>, <p>, <div>, <li> etc. with newlines, other tags with space
        text = re.sub(r'<br\s*/?>|</p>|</div>|</li>|</tr>', '\n', html_text, flags=re.IGNORECASE)
        text = re.sub(r'<[^>]+>', ' ', text)
        text = unescape(text)
        # Collapse multiple spaces within each line, but preserve newlines
        lines = [re.sub(r'[ \t]+', ' ', line).strip() for line in text.split('\n')]
        return '\n'.join(line for line in lines if line)
    text = re.sub(r'<[^>]+>', ' ', html_text)
    text = unescape(text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _vtex_match_product(products: list, sku: str, config: dict) -> dict | None:
    """Match a product from VTEX results by SKU reference."""
    sku_lower = sku.lower()
    for p in products:
        ref = (p.get("productReference") or "").lower()
        if ref.startswith(sku_lower) or sku_lower in ref:
            p["_vtex_config"] = config
            p["_search_api"] = "intelligent"
            return p
        # Also check item-level refIds
        for item in (p.get("items") or []):
            item_ref = (item.get("referenceId") or [{}])[0].get("Value", "") if item.get("referenceId") else ""
            item_name = (item.get("name") or "").lower()
            if sku_lower in item_ref.lower() or sku_lower in item_name:
                p["_vtex_config"] = config
                p["_search_api"] = "intelligent"
                return p
    return None


def _vtex_search_product(sku: str, brand: str, product_name: str = "") -> dict | None:
    """Search for a product by SKU on VTEX using multiple strategies."""
    config = _get_vtex_config(brand)
    if not config:
        log.warning(f"No VTEX config for brand '{brand}'")
        return None

    domain = config["domain"]
    headers = {"Accept": "application/json", "User-Agent": "CertAPI/2.0"}
    is_numeric = sku.isdigit()

    # 1. Primary: Intelligent Search API with SKU
    try:
        url = f"https://{domain}/api/io/_v/api/intelligent-search/product_search/"
        resp = requests.get(url, params={"query": sku, "count": 5}, headers=headers, timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            products = data.get("products", [])
            if products:
                match = _vtex_match_product(products, sku, config)
                if match:
                    return match
                # If no ref match but results exist, return first (only for non-numeric SKUs)
                if not is_numeric:
                    products[0]["_vtex_config"] = config
                    products[0]["_search_api"] = "intelligent"
                    return products[0]
    except requests.RequestException as e:
        log.warning(f"VTEX intelligent search ({sku}) failed: {e}")

    # 2. Fallback: Catalog System API with RefId
    try:
        url = f"https://{domain}/api/catalog_system/pub/products/search"
        resp = requests.get(url, params={"fq": f"alternateIds_RefId:{sku}"}, headers=headers, timeout=15)
        if resp.status_code in (200, 206):
            data = resp.json()
            if data:
                data[0]["_vtex_config"] = config
                data[0]["_search_api"] = "catalog"
                return data[0]
    except requests.RequestException as e:
        log.warning(f"VTEX catalog search ({sku}) failed: {e}")

    # 3. For numeric SKUs: try searching by product name as fallback
    # Puket SKUs are 9 digits but VTEX productRef is 12 digits, so name search helps
    if is_numeric and product_name:
        try:
            # Use first 3-4 meaningful words of the product name
            name_words = [w for w in product_name.split() if len(w) > 2][:4]
            name_query = " ".join(name_words)
            if name_query:
                url = f"https://{domain}/api/io/_v/api/intelligent-search/product_search/"
                resp = requests.get(url, params={"query": name_query, "count": 10}, headers=headers, timeout=15)
                if resp.status_code == 200:
                    data = resp.json()
                    products = data.get("products", [])
                    if products:
                        match = _vtex_match_product(products, sku, config)
                        if match:
                            return match
        except requests.RequestException as e:
            log.warning(f"VTEX name search ({product_name}) failed: {e}")

    return None


def _extract_cert_sentences(text: str) -> list[str]:
    """Extract only sentences/fragments that contain certification keywords from a block of text."""
    if not text:
        return []
    # Split by common delimiters: newlines, periods, semicolons, pipes
    fragments = re.split(r'[\n;|]+|(?<=\.)\s+', text)
    results = []
    for frag in fragments:
        frag = frag.strip().rstrip('.')
        if not frag or len(frag) < 5:
            continue
        frag_lower = frag.lower()
        if any(kw in frag_lower for kw in CERT_KEYWORDS):
            results.append(frag.strip())
    return results


# Spec field names that directly contain certification info (e.g. marketplace products)
CERT_SPEC_NAMES = [
    "certificação inmetro", "certificacao inmetro",
    "certificação", "certificacao",
    "registro inmetro", "selo inmetro",
    "homologação anatel", "homologacao anatel",
    "registro anvisa",
]


def _extract_cert_text_from_vtex(product_data: dict) -> str:
    """Extract all certification-related text from VTEX product data.

    Checks multiple locations where certification info can appear:
    - Puket: typically in description (last sentence) and complementName
    - Imaginarium own products: in description
    - Imaginarium marketplace: in specificationGroups > "Certificação Inmetro"
    """
    found_texts: list[str] = []
    api_type = product_data.get("_search_api", "intelligent")

    if api_type == "intelligent":
        # Intelligent Search API response format

        # 1. Check description — use keep_newlines to preserve structure
        desc = _strip_html(product_data.get("description", ""), keep_newlines=True)
        found_texts.extend(_extract_cert_sentences(desc))

        # 2. Check properties (dedicated cert fields)
        for prop in (product_data.get("properties") or []):
            if isinstance(prop, dict):
                name = prop.get("name", "")
                values = prop.get("values", [])
                name_lower = name.lower().strip()
                # Directly include values from known cert spec fields
                is_cert_field = any(cn in name_lower for cn in CERT_SPEC_NAMES)
                for val in values:
                    val_str = _strip_html(str(val))
                    if is_cert_field and val_str:
                        found_texts.append(f"{name}: {val_str}")
                    elif any(kw in val_str.lower() for kw in CERT_KEYWORDS) or any(kw in name_lower for kw in CERT_KEYWORDS):
                        found_texts.append(f"{name}: {val_str}")

        # 3. Check specificationGroups (key for marketplace products in "Detalhes")
        for group in (product_data.get("specificationGroups") or []):
            if isinstance(group, dict):
                for spec in (group.get("specifications") or []):
                    if isinstance(spec, dict):
                        name = spec.get("name", "")
                        values = spec.get("values", [])
                        name_lower = name.lower().strip()
                        is_cert_field = any(cn in name_lower for cn in CERT_SPEC_NAMES)
                        for val in values:
                            val_str = _strip_html(str(val))
                            if is_cert_field and val_str:
                                found_texts.append(f"{name}: {val_str}")
                            elif any(kw in val_str.lower() for kw in CERT_KEYWORDS) or any(kw in name_lower for kw in CERT_KEYWORDS):
                                found_texts.append(f"{name}: {val_str}")

        # 4. Check items (SKU-level) — complementName is key for certification text
        for item in (product_data.get("items") or []):
            if isinstance(item, dict):
                comp_raw = item.get("complementName", "")
                if not comp_raw:
                    continue
                # Use keep_newlines to preserve line structure in complementName
                comp = _strip_html(comp_raw, keep_newlines=True)
                cert_sentences = _extract_cert_sentences(comp)
                if cert_sentences:
                    found_texts.extend(cert_sentences)

    else:
        # Catalog System API response format (legacy fallback)
        desc = _strip_html(product_data.get("description", ""), keep_newlines=True)
        found_texts.extend(_extract_cert_sentences(desc))

        all_specs = product_data.get("allSpecifications", [])
        for spec_name in all_specs:
            spec_values = product_data.get(spec_name, [])
            name_lower = spec_name.lower().strip()
            is_cert_field = any(cn in name_lower for cn in CERT_SPEC_NAMES)
            if isinstance(spec_values, list):
                for val in spec_values:
                    val_str = _strip_html(str(val))
                    if is_cert_field and val_str:
                        found_texts.append(f"{spec_name}: {val_str}")
                    elif any(kw in val_str.lower() for kw in CERT_KEYWORDS) or any(kw in name_lower for kw in CERT_KEYWORDS):
                        found_texts.append(f"{spec_name}: {val_str}")

        for item in (product_data.get("items") or []):
            comp_raw = item.get("complementName", "")
            if not comp_raw:
                continue
            comp = _strip_html(comp_raw, keep_newlines=True)
            cert_sentences = _extract_cert_sentences(comp)
            if cert_sentences:
                found_texts.extend(cert_sentences)

    # Deduplicate while preserving order
    seen = set()
    unique = []
    for t in found_texts:
        t_clean = t.strip()
        if t_clean and t_clean not in seen:
            seen.add(t_clean)
            unique.append(t_clean)

    return " | ".join(unique) if unique else ""


def _normalize_reg_number(num: str) -> str:
    """Normalize registration numbers by stripping leading zeros.
    '006083/2024' -> '6083/2024', '0098' -> '98'
    """
    parts = num.split('/')
    return '/'.join(p.lstrip('0') or '0' for p in parts)


def _build_product_url(product_data: dict) -> str:
    """Build the product page URL from VTEX data."""
    config = product_data.get("_vtex_config", {})
    site_url = config.get("site_url", "")
    link = product_data.get("link", "")

    if link and link.startswith("http"):
        return link
    if link and site_url:
        return f"{site_url}{link}"
    link_text = product_data.get("linkText", "")
    if link_text and site_url:
        return f"{site_url}/{link_text}/p"
    return ""


def _extract_cert_body(text: str) -> set[str]:
    """Identify which certification bodies are mentioned in a text."""
    bodies = set()
    lower = text.lower()
    for body in ("inmetro", "anvisa", "anatel", "abnt"):
        if body in lower:
            bodies.add(body)
    # OCP / BRICS are INMETRO-accredited certification bodies
    # "Produto certificado por BRICS OCP 0098" → INMETRO
    if "inmetro" not in bodies and ("ocp " in lower or "brics" in lower):
        bodies.add("inmetro")
    return bodies


def _has_registration_number(text: str) -> bool:
    """Check if text contains a registration/certificate number pattern."""
    # Patterns: 006083/2024, Nº 12345, Registro 006083, OCP 0098, Anatel: 07388-24-15956, etc.
    patterns = [
        r'\d{4,}/\d{4}',                       # 006083/2024
        r'n[°ºo]\s*\.?\s*\d{3,}',              # Nº 006083
        r'registro\s+\d{3,}',                   # Registro 006083
        r'ocp\s+\d{3,}',                       # OCP 0098
        r'certificado\s+(?:n[°ºo]?\s*)?\.?\s*\d{3,}',  # Certificado Nº 12345
        r'homologa[çc][ãa]o[^:]*:\s*\d{3,}',   # homologação Anatel: 07388-24-15956
        r'\d{4,}-\d{2}-\d{4,}',                # 07388-24-15956 (Anatel code format)
        r'ce-\w+[\s/]\w+\s+\d{3,}',            # CE-BRI/BRICS 01180-20 (certificate number)
    ]
    lower = text.lower()
    return any(re.search(p, lower) for p in patterns)


def _compare_ecommerce_description(ecommerce_desc: str, actual: str) -> tuple[str, float] | None:
    """
    Compare the exact e-commerce description (from the spreadsheet) against
    the actual text found on the site. Returns (status, score) or None if
    no e-commerce description is available.

    The e-commerce description is the EXACT text that should appear on the
    product page (e.g. "Registro Inmetro 010208/2024" or
    "CE-BRI/ICEPEX-N 01264-25").
    """
    if not ecommerce_desc:
        return None
    if not actual:
        return ("URL_NOT_FOUND", 0.0)

    desc_clean = ecommerce_desc.strip().lower()
    actual_clean = actual.strip().lower()

    # Exact substring match — the e-commerce description should appear in the site text
    if desc_clean in actual_clean:
        return ("OK", 1.0)

    # Try matching key fragments (registration numbers, cert codes)
    # Extract numbers/codes from the expected description
    codes = re.findall(r'[\w-]{4,}/[\w-]+|[\w-]+-[\w-]+-[\w-]+|\d{5,}', ecommerce_desc)
    if codes:
        # Normalize codes (strip leading zeros) before comparing
        normalized_codes = [_normalize_reg_number(c) for c in codes]
        found = sum(
            1 for nc in normalized_codes
            if nc.lower() in actual_clean or any(_normalize_reg_number(a) == nc.lower() for a in re.findall(r'[\w-]{4,}/[\w-]+|[\w-]+-[\w-]+-[\w-]+|\d{5,}', actual))
        )
        if found == len(codes):
            return ("OK", 0.95)
        if found > 0:
            return ("INCONSISTENT", 0.5 + 0.3 * (found / len(codes)))

    # Sequence similarity as last resort
    seq_score = difflib.SequenceMatcher(None, desc_clean, actual_clean).ratio()
    if seq_score >= 0.8:
        return ("OK", seq_score)
    if seq_score >= 0.4:
        return ("INCONSISTENT", seq_score)

    return ("URL_NOT_FOUND", seq_score)


def _compare_cert_texts(expected: str, actual: str, ecommerce_desc: str = "") -> tuple[str, float]:
    """
    Compare expected certification info with actual text found on site.

    When 'ecommerce_desc' (Descrição E-commerce from the spreadsheet) is available,
    uses exact substring matching first — this is the most reliable comparison since
    it's the exact text that should appear on the product page.

    Falls back to cert body + registration number matching using 'expected'
    (certification_type from the spreadsheet).

    Returns (status, score).
    """
    if not expected and not ecommerce_desc:
        return ("NO_EXPECTED", 0.0)
    if not actual:
        return ("URL_NOT_FOUND", 0.0)

    # Priority 1: Compare against e-commerce description (exact text match)
    ecom_result = _compare_ecommerce_description(ecommerce_desc, actual)
    if ecom_result is not None:
        return ecom_result

    if not expected:
        return ("NO_EXPECTED", 0.0)

    exp_lower = expected.lower().strip()
    act_lower = actual.lower().strip()

    # 0. Handle "ENCERRAMENTO" products — these are being phased out
    # If expected says "ENCERRAMENTO", the product is in phase-out.
    # If the site still has valid certification, it's OK with full score (cert still valid during phase-out).
    # If no cert on site, it's OK too (cert may have been removed as part of phase-out).
    if 'encerramento' in exp_lower:
        if actual and (_extract_cert_body(actual) or _has_registration_number(actual)):
            return ("OK", 1.0)  # Still has valid cert during phase-out
        return ("OK", 1.0)  # Phase-out product, no cert needed

    # 1. Check if expected has specific registration numbers that should appear on site
    exp_reg_numbers = re.findall(r'\d{4,}/\d{4}', expected)
    if exp_reg_numbers:
        # Also extract reg numbers from actual and normalize both sides
        act_reg_numbers = re.findall(r'\d{4,}/\d{4}', actual)
        norm_act = {_normalize_reg_number(n) for n in act_reg_numbers}
        found_count = sum(1 for num in exp_reg_numbers if num in actual or _normalize_reg_number(num) in norm_act)
        if found_count == len(exp_reg_numbers):
            return ("OK", 1.0)

    # 2. Check certification body match (most important check)
    exp_bodies = _extract_cert_body(expected)
    act_bodies = _extract_cert_body(actual)
    matching_bodies = exp_bodies & act_bodies
    actual_has_reg = _has_registration_number(actual)

    # If same cert body AND the site shows a registration number → OK
    # This is the typical case: expected="INMETRO BRINQUEDOS ...", actual="Certificação INMETRO: ... Nº Registro 006083/2024"
    if matching_bodies and actual_has_reg:
        return ("OK", 0.95)

    # If same cert body but no registration number → INCONSISTENT (cert mentioned but incomplete)
    if matching_bodies and not actual_has_reg:
        return ("INCONSISTENT", 0.6)

    # If expected doesn't mention any cert body (just a product type like "ESTOJO")
    # but actual has a cert body + registration number → product is certified
    if not exp_bodies and act_bodies and actual_has_reg:
        return ("OK", 0.9)

    # 3. Check for portaria numbers
    portaria_expected = re.findall(r'portaria\s*(?:n[°º.]?\s*)?(\d+)', exp_lower)
    if portaria_expected:
        portaria_found = all(
            re.search(rf'(?:portaria|port\.?)\s*(?:n[°º.]?\s*)?{num}', act_lower)
            for num in portaria_expected
        )
        if portaria_found:
            return ("OK", 0.9)

    # 4. Word overlap + sequence similarity for other cases
    exp_words = set(w for w in re.findall(r'\w+', exp_lower) if len(w) > 3)
    act_words = set(w for w in re.findall(r'\w+', act_lower) if len(w) > 3)

    if exp_words:
        overlap = exp_words & act_words
        word_score = len(overlap) / len(exp_words)
    else:
        word_score = 0.0

    seq_score = difflib.SequenceMatcher(None, exp_lower, act_lower).ratio()
    score = max(word_score, seq_score)

    # If actual has a registration number from ANY cert body, slight boost
    if actual_has_reg and score >= 0.2:
        score = max(score, 0.5)

    # Determine status
    if score >= 0.7:
        return ("OK", score)
    elif score >= 0.3:
        return ("INCONSISTENT", score)
    else:
        return ("URL_NOT_FOUND", score)


def validate_single_product(
    sku: str, brand: str, expected_cert: str,
    product_name: str = "", ecommerce_description: str = "",
    sheet_status: str = "", is_expired: bool = False, sale_deadline_date: str = "",
) -> dict:
    """
    Validate a single product against VTEX.
    Returns dict with: status, score, url, actual_cert_text, error.
    """
    now = datetime.now(timezone.utc)
    result = {
        "sku": sku,
        "brand": brand,
        "status": "URL_NOT_FOUND",
        "score": 0.0,
        "url": None,
        "actual_cert_text": None,
        "expected_cert_text": expected_cert or None,
        "ecommerce_description": ecommerce_description or None,
        "error": None,
        "verified_at": now.isoformat(),
    }

    # Early-exit for expired products — no need to hit VTEX
    if is_expired and sale_deadline_date:
        result["status"] = "EXPIRED"
        result["error"] = f"Certificado vencido - prazo final venda: {sale_deadline_date}"
        return result

    if not expected_cert and not ecommerce_description:
        result["status"] = "NO_EXPECTED"
        return result

    config = _get_vtex_config(brand)
    if not config:
        result["status"] = "API_ERROR"
        result["error"] = f"No VTEX store configured for brand '{brand}'"
        return result

    try:
        vtex_product = _vtex_search_product(sku, brand, product_name=product_name)
    except Exception as e:
        result["status"] = "API_ERROR"
        result["error"] = f"VTEX API error: {str(e)}"
        return result

    if not vtex_product:
        result["status"] = "URL_NOT_FOUND"
        result["error"] = f"Product SKU {sku} not found on {config['domain']}"
        return result

    # Build product URL
    result["url"] = _build_product_url(vtex_product)

    # Extract certification text from VTEX product
    actual_cert = _extract_cert_text_from_vtex(vtex_product)
    result["actual_cert_text"] = actual_cert if actual_cert else None

    # Compare expected vs actual
    status, score = _compare_cert_texts(expected_cert, actual_cert, ecommerce_description)
    result["status"] = status
    result["score"] = round(score, 3)

    if status == "URL_NOT_FOUND" and result["url"]:
        result["error"] = "No certification text found on product page"
    elif status == "INCONSISTENT":
        result["error"] = "Certification text found but does not match expected"

    return result


# ---------------------------------------------------------------------------
# Startup
# ---------------------------------------------------------------------------

scheduler = BackgroundScheduler(timezone="America/Sao_Paulo")


def _execute_schedule(schedule_id: str, brand_filter: str | None):
    """Execute a scheduled validation run."""
    log.info(f"Scheduler executing schedule {schedule_id} (brand={brand_filter})")
    try:
        run_id = str(uuid.uuid4())
        with db() as (conn, cur):
            cur.execute(
                "INSERT INTO cert_validation_runs (id, status, brand_filter) VALUES (%s, 'running', %s)",
                [run_id, brand_filter],
            )
            now = datetime.now(timezone.utc)
            cur.execute("UPDATE cert_schedules SET last_run = %s WHERE id = %s", [now, schedule_id])
            cur.execute(
                "INSERT INTO cert_schedule_history (schedule_id, status) VALUES (%s, 'running')",
                [schedule_id],
            )

        _running_validations[run_id] = {"status": "running", "events": [], "processed": 0, "total": 0, "_started_at": time.time()}
        # Run synchronously in the scheduler thread (APScheduler manages threading)
        _run_validation(run_id, brand_filter, None, "sheets")

        # Update history to completed
        with db() as (conn, cur):
            state = _running_validations.get(run_id, {})
            summary = {
                "total": state.get("total", 0),
                "ok": sum(1 for e in state.get("events", []) if e.get("product", {}).get("status") == "OK"),
                "inconsistent": sum(1 for e in state.get("events", []) if e.get("product", {}).get("status") == "INCONSISTENT"),
                "not_found": sum(1 for e in state.get("events", []) if e.get("product", {}).get("status") not in ("OK", "INCONSISTENT")),
            }
            cur.execute(
                """UPDATE cert_schedule_history SET status = 'completed', summary = %s
                   WHERE id = (
                     SELECT id FROM cert_schedule_history
                     WHERE schedule_id = %s AND status = 'running'
                     ORDER BY run_date DESC LIMIT 1
                   )""",
                [json.dumps(summary), schedule_id],
            )
        log.info(f"Schedule {schedule_id} completed: {summary}")
    except Exception as e:
        log.error(f"Schedule {schedule_id} failed: {e}")
        try:
            with db() as (conn, cur):
                cur.execute(
                    """UPDATE cert_schedule_history SET status = 'failed'
                       WHERE id = (
                         SELECT id FROM cert_schedule_history
                         WHERE schedule_id = %s AND status = 'running'
                         ORDER BY run_date DESC LIMIT 1
                       )""",
                    [schedule_id],
                )
        except Exception:
            pass


def _load_schedules_into_scheduler():
    """Load all enabled schedules from DB into APScheduler."""
    if not DATABASE_URL:
        return
    try:
        # Remove all existing schedule jobs
        for job in scheduler.get_jobs():
            if job.id.startswith("cert_schedule_"):
                job.remove()

        with db() as (conn, cur):
            cur.execute("SELECT * FROM cert_schedules WHERE enabled = true")
            schedules = [dict(r) for r in cur.fetchall()]

        def _validate_cron_part(part: str) -> bool:
            return bool(re.match(r'^[\d,\-\*/]+$', part))

        for s in schedules:
            cron_expr = s["cron_expression"]
            parts = cron_expr.split()
            if len(parts) != 5 or not all(_validate_cron_part(p) for p in parts):
                log.warning(f"Invalid cron for schedule {s['id']}: {cron_expr}")
                continue

            try:
                trigger = CronTrigger(
                    minute=parts[0],
                    hour=parts[1],
                    day=parts[2],
                    month=parts[3],
                    day_of_week=parts[4],
                    timezone="America/Sao_Paulo",
                )
                job_id = f"cert_schedule_{s['id']}"
                scheduler.add_job(
                    _execute_schedule,
                    trigger=trigger,
                    id=job_id,
                    args=[str(s["id"]), s.get("brand_filter")],
                    replace_existing=True,
                    max_instances=1,
                )
                # Calculate and store next_run
                next_run = trigger.get_next_fire_time(None, datetime.now(timezone.utc))
                if next_run:
                    with db() as (conn, cur):
                        cur.execute("UPDATE cert_schedules SET next_run = %s WHERE id = %s", [next_run, s["id"]])
                log.info(f"Loaded schedule '{s['name']}' ({cron_expr}) next_run={next_run}")
            except Exception as e:
                log.warning(f"Failed to load schedule {s['id']}: {e}")

    except Exception as e:
        log.error(f"Failed to load schedules: {e}")


@app.on_event("startup")
def startup():
    if DATABASE_URL:
        try:
            ensure_tables()
        except Exception as e:
            log.warning(f"Could not create tables: {e}")

    if SHEETS_CLIENT_EMAIL and SHEETS_PRIVATE_KEY:
        try:
            result = sync_sheets_to_db()
            log.info(f"Startup sheets sync: {result}")
        except Exception as e:
            log.warning(f"Startup sheets sync failed: {e}")
        try:
            li_result = sync_licenciados_to_db()
            log.info(f"Startup licenciados sync: {li_result}")
        except Exception as e:
            log.warning(f"Startup licenciados sync failed: {e}")

    # Start the APScheduler and load cron schedules from DB
    try:
        scheduler.start()
        _load_schedules_into_scheduler()
        log.info("APScheduler started successfully")
    except Exception as e:
        log.warning(f"Failed to start scheduler: {e}")


@app.on_event("shutdown")
def shutdown():
    try:
        scheduler.shutdown(wait=False)
    except Exception:
        pass
    if _pool and not _pool.closed:
        _pool.closeall()


# ---------------------------------------------------------------------------
# Health
# ---------------------------------------------------------------------------

@app.get("/api/health")
def health():
    result = {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}
    if DATABASE_URL:
        try:
            with db() as (conn, cur):
                cur.execute("SELECT 1")
            result["database"] = "connected"
        except Exception:
            result["database"] = "disconnected"
    result["sheets_configured"] = bool(SHEETS_CLIENT_EMAIL and SHEETS_PRIVATE_KEY)
    result["vtex_stores"] = list(VTEX_STORES.keys())
    return result


# ---------------------------------------------------------------------------
# Sheets sync
# ---------------------------------------------------------------------------

@app.post("/api/sync-sheets")
@limiter.limit("5/minute")
def trigger_sync_sheets(request: Request):
    if not SHEETS_CLIENT_EMAIL or not SHEETS_PRIVATE_KEY:
        raise HTTPException(400, "Google Sheets credentials not configured")
    result = sync_sheets_to_db()
    if result.get("error"):
        raise HTTPException(500, result["error"])
    return result


# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------

@app.get("/api/stats")
def get_stats():
    if not DATABASE_URL:
        return _empty_stats()
    try:
        with db() as (conn, cur):
            cur.execute("SELECT COUNT(*) as cnt FROM cert_products")
            total = cur.fetchone()["cnt"]

            cur.execute("""
                SELECT * FROM cert_validation_runs
                WHERE status = 'completed'
                ORDER BY finished_at DESC NULLS LAST
                LIMIT 1
            """)
            last_run_row = cur.fetchone()

            last_run = None
            if last_run_row:
                last_run = {
                    "date": (last_run_row["finished_at"] or last_run_row["started_at"]).isoformat(),
                    "total": last_run_row["total"],
                    "ok": last_run_row["ok"],
                    "inconsistent": last_run_row["inconsistent"],
                    "not_found": (last_run_row.get("not_found") or 0) + (last_run_row.get("missing") or 0),
                }

            cur.execute("""
                SELECT brand,
                    COUNT(*) FILTER (WHERE last_validation_status = 'OK') as ok,
                    COUNT(*) FILTER (WHERE last_validation_status = 'INCONSISTENT') as inconsistent,
                    COUNT(*) FILTER (WHERE last_validation_status NOT IN ('OK','INCONSISTENT') OR last_validation_status IS NULL) as not_found,
                    COUNT(*) FILTER (WHERE is_expired = TRUE) as expired
                FROM cert_products
                WHERE brand != ''
                GROUP BY brand
                ORDER BY brand
            """)
            by_brand = [dict(r) for r in cur.fetchall()]

            # Count total expired products
            cur.execute("SELECT COUNT(*) as cnt FROM cert_products WHERE is_expired = TRUE")
            total_expired = cur.fetchone()["cnt"]

            return {
                "total_products": total,
                "total_expired": total_expired,
                "last_run": last_run,
                "by_brand": by_brand,
            }
    except Exception:
        return _empty_stats()


def _empty_stats():
    return {"total_products": 0, "total_expired": 0, "last_run": None, "by_brand": []}


# ---------------------------------------------------------------------------
# Expired products (Encerramentos)
# ---------------------------------------------------------------------------

@app.get("/api/expired")
def list_expired_products(
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=100),
    search: str = Query(""),
    brand: str = Query(""),
):
    if not DATABASE_URL:
        return {"products": [], "total": 0, "page": 1, "per_page": per_page, "total_pages": 0}

    with db() as (conn, cur):
        conditions = ["is_expired = TRUE"]
        params: list = []

        if search:
            conditions.append("(sku ILIKE %s OR name ILIKE %s)")
            params.extend([f"%{search}%", f"%{search}%"])
        if brand:
            conditions.append("LOWER(brand) = LOWER(%s)")
            params.append(brand)

        where = "WHERE " + " AND ".join(conditions)

        cur.execute(f"SELECT COUNT(*) as cnt FROM cert_products {where}", params)
        total = cur.fetchone()["cnt"]

        offset = (page - 1) * per_page
        cur.execute(
            f"SELECT * FROM cert_products {where} ORDER BY sale_deadline_date ASC NULLS LAST LIMIT %s OFFSET %s",
            params + [per_page, offset],
        )
        rows = [_serialize_product(dict(r)) for r in cur.fetchall()]
        total_pages = max(1, (total + per_page - 1) // per_page)

        return {
            "products": rows,
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": total_pages,
        }


# ---------------------------------------------------------------------------
# Products
# ---------------------------------------------------------------------------

def _serialize_product(r: dict) -> dict:
    for dtfield in ("last_validation_date", "created_at", "updated_at", "sale_deadline_date"):
        if r.get(dtfield):
            r[dtfield] = r[dtfield].isoformat() if hasattr(r[dtfield], 'isoformat') else str(r[dtfield])
    return r


@app.get("/api/products")
def list_products(
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=100),
    search: str = Query(""),
    brand: str = Query(""),
    status: str = Query(""),
    start_date: str = Query(""),
    end_date: str = Query(""),
):
    if not DATABASE_URL:
        return {"products": [], "total": 0, "page": 1, "per_page": per_page, "total_pages": 0, "last_validation_date": None}

    with db() as (conn, cur):
        conditions = []
        params: list = []

        if search:
            conditions.append("(sku ILIKE %s OR name ILIKE %s)")
            params.extend([f"%{search}%", f"%{search}%"])
        if brand:
            conditions.append("LOWER(brand) = LOWER(%s)")
            params.append(brand)
        if status:
            statuses = [s.strip() for s in status.split(",") if s.strip()]
            if "EXPIRED" in statuses:
                statuses.remove("EXPIRED")
                if statuses:
                    conditions.append("(last_validation_status IN ({}) OR is_expired = TRUE)".format(",".join(["%s"] * len(statuses))))
                    params.extend(statuses)
                else:
                    conditions.append("is_expired = TRUE")
            elif statuses:
                conditions.append("last_validation_status IN ({})".format(",".join(["%s"] * len(statuses))))
                params.extend(statuses)

        if start_date:
            conditions.append("last_validation_date >= %s::date")
            params.append(start_date)
        if end_date:
            conditions.append("last_validation_date < (%s::date + interval '1 day')")
            params.append(end_date)

        where = "WHERE " + " AND ".join(conditions) if conditions else ""

        cur.execute(f"SELECT COUNT(*) as cnt FROM cert_products {where}", params)
        total = cur.fetchone()["cnt"]

        offset = (page - 1) * per_page
        cur.execute(
            f"SELECT * FROM cert_products {where} ORDER BY sku LIMIT %s OFFSET %s",
            params + [per_page, offset],
        )
        products_raw = [_serialize_product(dict(r)) for r in cur.fetchall()]

        # Enrich with stock data
        if products_raw:
            skus = [p["sku"] for p in products_raw]
            placeholders = ",".join(["%s"] * len(skus))
            cur.execute(f"""
                SELECT sku, source, warehouse,
                    COALESCE(SUM(quantity), 0) as qty,
                    COALESCE(SUM(available), 0) as avail
                FROM cert_stock WHERE sku IN ({placeholders})
                GROUP BY sku, source, warehouse
            """, skus)
            stock_rows = cur.fetchall()

            stock_map: dict = {}
            for sr in stock_rows:
                sk = sr["sku"]
                if sk not in stock_map:
                    stock_map[sk] = {"stock_cd": 0, "stock_ecommerce": 0, "stock_total": 0, "stock_detail": []}
                entry = {"source": sr["source"], "warehouse": sr["warehouse"], "quantity": sr["qty"], "available": sr["avail"]}
                stock_map[sk]["stock_detail"].append(entry)
                if sr["source"] == "wms_biguacu":
                    stock_map[sk]["stock_cd"] += sr["avail"] or sr["qty"] or 0
                else:
                    stock_map[sk]["stock_ecommerce"] += sr["avail"] or sr["qty"] or 0
                stock_map[sk]["stock_total"] = stock_map[sk]["stock_cd"] + stock_map[sk]["stock_ecommerce"]

            for p in products_raw:
                s = stock_map.get(p["sku"], {})
                p["stock_cd"] = s.get("stock_cd", 0)
                p["stock_ecommerce"] = s.get("stock_ecommerce", 0)
                p["stock_total"] = s.get("stock_total", 0)
                p["stock_detail"] = s.get("stock_detail", [])

        rows = products_raw

        cur.execute("SELECT MAX(last_validation_date) as last_date FROM cert_products")
        last_date_row = cur.fetchone()
        last_date = last_date_row["last_date"].isoformat() if last_date_row and last_date_row["last_date"] else None

        total_pages = max(1, (total + per_page - 1) // per_page)

        return {
            "products": rows,
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": total_pages,
            "last_validation_date": last_date,
        }


@app.get("/api/products/{sku}")
def get_product(sku: str):
    if not DATABASE_URL:
        raise HTTPException(404, "Product not found")

    with db() as (conn, cur):
        cur.execute("SELECT * FROM cert_products WHERE sku = %s", [sku])
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Product not found")
        result = _serialize_product(dict(row))

        # Build last_validation object for frontend detail page
        if result.get("last_validation_status"):
            result["last_validation"] = {
                "status": result["last_validation_status"],
                "score": result.get("last_validation_score"),
                "url": result.get("last_validation_url"),
                "actual_cert_text": result.get("actual_cert_text"),
                "error": result.get("last_validation_error"),
                "date": result.get("last_validation_date"),
            }
        return result


class VerifyRequest(BaseModel):
    sku: str
    brand: str


@app.post("/api/products/verify")
def verify_product(req: VerifyRequest):
    """Verify a single product against the VTEX e-commerce site in real time."""
    now = datetime.now(timezone.utc)

    # Get expected cert text, product name, and ecommerce description from DB
    expected_cert = ""
    product_name = ""
    ecommerce_desc = ""
    sheet_status = ""
    is_expired = False
    sale_deadline_date_str = ""
    if DATABASE_URL:
        try:
            with db() as (conn, cur):
                cur.execute(
                    "SELECT certification_type, brand, name, ecommerce_description, sheet_status, is_expired, sale_deadline_date FROM cert_products WHERE sku = %s",
                    [req.sku],
                )
                row = cur.fetchone()
                if row:
                    expected_cert = row.get("certification_type", "") or ""
                    product_name = row.get("name", "") or ""
                    ecommerce_desc = row.get("ecommerce_description", "") or ""
                    sheet_status = row.get("sheet_status", "") or ""
                    is_expired = bool(row.get("is_expired", False))
                    sale_deadline_date_str = str(row["sale_deadline_date"]) if row.get("sale_deadline_date") else ""
        except Exception:
            pass

    # Real VTEX validation
    result = validate_single_product(
        req.sku, req.brand, expected_cert,
        product_name=product_name, ecommerce_description=ecommerce_desc,
        sheet_status=sheet_status, is_expired=is_expired, sale_deadline_date=sale_deadline_date_str,
    )

    # Save result to DB
    if DATABASE_URL:
        try:
            with db() as (conn, cur):
                cur.execute(
                    """
                    UPDATE cert_products
                    SET last_validation_status = %s,
                        last_validation_score = %s,
                        last_validation_url = %s,
                        last_validation_date = %s,
                        last_validation_error = %s,
                        actual_cert_text = %s,
                        updated_at = %s
                    WHERE sku = %s
                    """,
                    [result["status"], result["score"], result["url"],
                     now, result["error"], result.get("actual_cert_text"),
                     now, req.sku],
                )
        except Exception as e:
            log.error(f"Failed to update product {req.sku}: {e}")

    return result


# ---------------------------------------------------------------------------
# Validation runs
# ---------------------------------------------------------------------------

class ValidateRequest(BaseModel):
    brand: str | None = None
    limit: int | None = None
    source: str | None = None


@app.post("/api/validate")
@limiter.limit("5/minute")
def start_validation(request: Request, req: ValidateRequest):
    run_id = str(uuid.uuid4())

    if DATABASE_URL:
        with db() as (conn, cur):
            cur.execute(
                "INSERT INTO cert_validation_runs (id, status, brand_filter) VALUES (%s, 'running', %s)",
                [run_id, req.brand],
            )

    _cleanup_old_validations()
    _running_validations[run_id] = {"status": "running", "events": [], "processed": 0, "total": 0, "_started_at": time.time()}
    thread = threading.Thread(
        target=_run_validation, args=(run_id, req.brand, req.limit, req.source), daemon=True
    )
    thread.start()

    return {"run_id": run_id, "status": "running"}


def _run_validation(run_id: str, brand_filter: str | None, limit: int | None, source: str | None = None):
    """Background validation — queries VTEX for each product."""
    state = _running_validations[run_id]
    try:
        # Sync from sheets if requested
        if source == "sheets" and SHEETS_CLIENT_EMAIL and SHEETS_PRIVATE_KEY:
            try:
                sync_result = sync_sheets_to_db()
                log.info(f"Pre-validation sheets sync: {sync_result}")
            except Exception as e:
                log.warning(f"Pre-validation sheets sync failed: {e}")

        products = []
        if DATABASE_URL:
            with db() as (conn, cur):
                conditions = []
                params: list = []
                if brand_filter:
                    conditions.append("LOWER(brand) = LOWER(%s)")
                    params.append(brand_filter)
                where = "WHERE " + " AND ".join(conditions) if conditions else ""
                sql = f"SELECT sku, name, brand, certification_type, ecommerce_description, sheet_status, is_expired, sale_deadline_date FROM cert_products {where} ORDER BY sku"
                if limit:
                    sql += f" LIMIT {int(limit)}"
                cur.execute(sql, params)
                products = [dict(r) for r in cur.fetchall()]

        state["total"] = len(products)
        ok = inconsistent = not_found = 0
        now = datetime.now(timezone.utc)
        report_products = []

        for i, p in enumerate(products):
            sku = p["sku"]
            brand = p["brand"]
            expected_cert = p.get("certification_type", "") or ""

            # Real VTEX validation
            ecommerce_desc = p.get("ecommerce_description", "") or ""
            p_is_expired = bool(p.get("is_expired", False))
            p_sale_deadline = str(p["sale_deadline_date"]) if p.get("sale_deadline_date") else ""
            vresult = validate_single_product(
                sku, brand, expected_cert,
                product_name=p.get("name", ""), ecommerce_description=ecommerce_desc,
                sheet_status=p.get("sheet_status", "") or "",
                is_expired=p_is_expired, sale_deadline_date=p_sale_deadline,
            )

            status = vresult["status"]
            score = vresult["score"]

            if status == "OK":
                ok += 1
            elif status == "INCONSISTENT":
                inconsistent += 1
            else:
                not_found += 1

            state["processed"] = i + 1
            event = {
                "type": "progress",
                "current": i + 1,
                "total": len(products),
                "product": {"sku": sku, "name": p["name"], "status": status, "score": score},
            }
            state["events"].append(event)

            report_products.append({
                "sku": sku,
                "name": p["name"],
                "brand": brand,
                "status": status,
                "score": score,
                "url": vresult.get("url"),
                "actual_cert_text": vresult.get("actual_cert_text"),
                "expected_cert_text": expected_cert,
                "error": vresult.get("error"),
            })

            # Update product in DB
            if DATABASE_URL:
                try:
                    with db() as (conn, cur):
                        cur.execute(
                            """
                            UPDATE cert_products
                            SET last_validation_status = %s,
                                last_validation_score = %s,
                                last_validation_url = %s,
                                last_validation_date = %s,
                                last_validation_error = %s,
                                actual_cert_text = %s,
                                updated_at = %s
                            WHERE sku = %s
                            """,
                            [status, score, vresult.get("url"), now,
                             vresult.get("error"), vresult.get("actual_cert_text"),
                             now, sku],
                        )
                except Exception:
                    pass

            # Rate limit: delay between VTEX requests
            if i < len(products) - 1:
                time.sleep(VTEX_REQUEST_DELAY)

        summary = {
            "total": len(products),
            "ok": ok,
            "missing": 0,
            "inconsistent": inconsistent,
            "not_found": not_found,
        }

        # Save report
        report_filename = f"validation_{run_id[:8]}_{now.strftime('%Y%m%d_%H%M%S')}.json"
        report_path = REPORTS_DIR / report_filename
        report_data = {
            "run_id": run_id,
            "date": now.isoformat(),
            "summary": summary,
            "products": report_products,
        }
        report_path.write_text(json.dumps(report_data, indent=2, ensure_ascii=False))

        # Update run in DB
        if DATABASE_URL:
            try:
                with db() as (conn, cur):
                    cur.execute(
                        """
                        UPDATE cert_validation_runs
                        SET status = 'completed', total = %s, ok = %s, missing = %s,
                            inconsistent = %s, not_found = %s, finished_at = %s, report_file = %s
                        WHERE id = %s
                        """,
                        [len(products), ok, missing, inconsistent, not_found,
                         datetime.now(timezone.utc), report_filename, run_id],
                    )
            except Exception:
                pass

        state["status"] = "completed"
        state["_finished_at"] = time.time()
        state["events"].append({"type": "complete", "summary": summary})

    except Exception as e:
        log.error(f"Validation run {run_id} failed: {e}", exc_info=True)
        state["status"] = "error"
        state["_finished_at"] = time.time()
        state["events"].append({"type": "error", "error": str(e)})


@app.get("/api/validate/{run_id}")
def get_validation_status(run_id: str):
    state = _running_validations.get(run_id)
    if state:
        return {
            "run_id": run_id,
            "status": state["status"],
            "processed": state["processed"],
            "total": state["total"],
        }

    if DATABASE_URL:
        with db() as (conn, cur):
            cur.execute("SELECT * FROM cert_validation_runs WHERE id = %s", [run_id])
            row = cur.fetchone()
            if row:
                return {
                    "run_id": run_id,
                    "status": row["status"],
                    "processed": row["total"],
                    "total": row["total"],
                }

    raise HTTPException(404, "Validation run not found")


@app.get("/api/validate/{run_id}/stream")
def stream_validation(run_id: str):
    """Server-Sent Events stream for validation progress."""
    state = _running_validations.get(run_id)
    if not state:
        raise HTTPException(404, "Validation run not found or already finished")

    def event_generator():
        sent = 0
        while True:
            events = state["events"]
            while sent < len(events):
                event = events[sent]
                yield f"data: {json.dumps(event)}\n\n"
                sent += 1
                if event.get("type") in ("complete", "error"):
                    return
            if state["status"] in ("completed", "error") and sent >= len(events):
                return
            time.sleep(0.3)

    return StreamingResponse(event_generator(), media_type="text/event-stream")


# ---------------------------------------------------------------------------
# Schedules
# ---------------------------------------------------------------------------

class ScheduleCreate(BaseModel):
    name: str
    cron: str
    brand_filter: str | None = None
    enabled: bool | None = True


class ScheduleUpdate(BaseModel):
    name: str | None = None
    cron: str | None = None
    brand_filter: str | None = None
    enabled: bool | None = None


@app.get("/api/schedules")
def list_schedules(
    start_date: str = Query(""),
    end_date: str = Query(""),
):
    if not DATABASE_URL:
        return []
    with db() as (conn, cur):
        conditions: list[str] = []
        params: list = []
        if start_date:
            conditions.append("last_run >= %s::date")
            params.append(start_date)
        if end_date:
            conditions.append("last_run < (%s::date + interval '1 day')")
            params.append(end_date)
        where = "WHERE " + " AND ".join(conditions) if conditions else ""
        cur.execute(f"SELECT * FROM cert_schedules {where} ORDER BY created_at DESC", params)
        rows = []
        for r in cur.fetchall():
            row = dict(r)
            for key in ("last_run", "next_run", "created_at"):
                if row.get(key):
                    row[key] = row[key].isoformat()
            row["id"] = str(row["id"])
            row["cron_expression"] = row.pop("cron_expression", "")
            rows.append(row)
        return rows


@app.post("/api/schedules")
def create_schedule(req: ScheduleCreate):
    if not DATABASE_URL:
        raise HTTPException(500, "Database not configured")
    with db() as (conn, cur):
        schedule_id = str(uuid.uuid4())
        cur.execute(
            """
            INSERT INTO cert_schedules (id, name, cron_expression, brand_filter, enabled)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING *
            """,
            [schedule_id, req.name, req.cron, req.brand_filter, req.enabled],
        )
        row = dict(cur.fetchone())
        for key in ("last_run", "next_run", "created_at"):
            if row.get(key):
                row[key] = row[key].isoformat()
        row["id"] = str(row["id"])
        row["cron_expression"] = row.pop("cron_expression", "")
        # Reload schedules into APScheduler
        _load_schedules_into_scheduler()
        return row


@app.put("/api/schedules/{schedule_id}")
def update_schedule(schedule_id: str, req: ScheduleUpdate):
    if not DATABASE_URL:
        raise HTTPException(500, "Database not configured")
    with db() as (conn, cur):
        updates = []
        params: list = []
        if req.name is not None:
            updates.append("name = %s")
            params.append(req.name)
        if req.cron is not None:
            updates.append("cron_expression = %s")
            params.append(req.cron)
        if req.brand_filter is not None:
            updates.append("brand_filter = %s")
            params.append(req.brand_filter if req.brand_filter else None)
        if req.enabled is not None:
            updates.append("enabled = %s")
            params.append(req.enabled)

        if not updates:
            raise HTTPException(400, "No fields to update")

        params.append(schedule_id)
        cur.execute(
            f"UPDATE cert_schedules SET {', '.join(updates)} WHERE id = %s RETURNING *",
            params,
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Schedule not found")
        result = dict(row)
        for key in ("last_run", "next_run", "created_at"):
            if result.get(key):
                result[key] = result[key].isoformat()
        result["id"] = str(result["id"])
        result["cron_expression"] = result.pop("cron_expression", "")
        # Reload schedules into APScheduler
        _load_schedules_into_scheduler()
        return result


@app.delete("/api/schedules/{schedule_id}")
def delete_schedule(schedule_id: str):
    if not DATABASE_URL:
        raise HTTPException(500, "Database not configured")
    with db() as (conn, cur):
        cur.execute("DELETE FROM cert_schedules WHERE id = %s", [schedule_id])
        if cur.rowcount == 0:
            raise HTTPException(404, "Schedule not found")
        # Reload schedules into APScheduler
        _load_schedules_into_scheduler()
        return {"ok": True}


@app.post("/api/schedules/{schedule_id}/run")
def run_schedule_now(schedule_id: str):
    if not DATABASE_URL:
        raise HTTPException(500, "Database not configured")
    with db() as (conn, cur):
        cur.execute("SELECT * FROM cert_schedules WHERE id = %s", [schedule_id])
        schedule = cur.fetchone()
        if not schedule:
            raise HTTPException(404, "Schedule not found")

        run_id = str(uuid.uuid4())
        brand = schedule["brand_filter"]
        cur.execute(
            "INSERT INTO cert_validation_runs (id, status, brand_filter) VALUES (%s, 'running', %s)",
            [run_id, brand],
        )
        now = datetime.now(timezone.utc)
        cur.execute("UPDATE cert_schedules SET last_run = %s WHERE id = %s", [now, schedule_id])

    _cleanup_old_validations()
    _running_validations[run_id] = {"status": "running", "events": [], "processed": 0, "total": 0, "_started_at": time.time()}
    thread = threading.Thread(target=_run_validation, args=(run_id, brand, None, "sheets"), daemon=True)
    thread.start()

    try:
        with db() as (conn, cur):
            cur.execute(
                "INSERT INTO cert_schedule_history (schedule_id, status) VALUES (%s, 'running')",
                [schedule_id],
            )
    except Exception:
        pass

    return {"run_id": run_id, "status": "running"}


@app.get("/api/schedules/{schedule_id}/history")
def get_schedule_history(schedule_id: str):
    if not DATABASE_URL:
        return []
    with db() as (conn, cur):
        cur.execute(
            """
            SELECT * FROM cert_schedule_history
            WHERE schedule_id = %s
            ORDER BY run_date DESC
            LIMIT 20
            """,
            [schedule_id],
        )
        rows = []
        for r in cur.fetchall():
            row = dict(r)
            row["id"] = str(row["id"])
            row["schedule_id"] = str(row["schedule_id"])
            if row.get("run_date"):
                row["run_date"] = row["run_date"].isoformat()
            rows.append(row)
        return rows


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------

@app.post("/api/reports/export")
@limiter.limit("10/minute")
def export_products_report(request: Request, brand: str = Query(""), status: str = Query("")):
    """Generate an Excel report from the current DB product data."""
    if not DATABASE_URL:
        raise HTTPException(500, "Database not configured")

    with db() as (conn, cur):
        conditions: list[str] = []
        params: list = []
        if brand:
            conditions.append("LOWER(brand) = LOWER(%s)")
            params.append(brand)
        if status:
            statuses = [s.strip() for s in status.split(",") if s.strip()]
            if "EXPIRED" in statuses:
                statuses.remove("EXPIRED")
                if statuses:
                    conditions.append("(last_validation_status IN ({}) OR is_expired = TRUE)".format(",".join(["%s"] * len(statuses))))
                    params.extend(statuses)
                else:
                    conditions.append("is_expired = TRUE")
            elif statuses:
                conditions.append("last_validation_status IN ({})".format(",".join(["%s"] * len(statuses))))
                params.extend(statuses)

        where = "WHERE " + " AND ".join(conditions) if conditions else ""
        cur.execute(f"SELECT * FROM cert_products {where} ORDER BY brand, sku", params)
        rows = [dict(r) for r in cur.fetchall()]

    now = datetime.now(timezone.utc)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    status_fills = {
        "OK": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "URL_NOT_FOUND": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "INCONSISTENT": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    }
    expired_fill = PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid")
    status_labels = {
        "OK": "Conforme", "INCONSISTENT": "Inconsistente",
        "URL_NOT_FOUND": "Nao Encontrado", "API_ERROR": "Erro de API",
        "NO_EXPECTED": "Sem Certificacao", "EXPIRED": "Vencido",
    }

    # Title
    ws.append(["Relatorio de Produtos - Certificacoes"])
    ws.merge_cells("A1:J1")
    ws["A1"].font = Font(bold=True, size=14, color="059669")
    ws.append([f"Gerado em: {now.strftime('%d/%m/%Y %H:%M')}"])
    ws.append([f"Total: {len(rows)} produtos"])
    ws.append([])

    # Count stats
    ok_count = sum(1 for r in rows if r.get("last_validation_status") == "OK")
    not_found_count = sum(1 for r in rows if r.get("last_validation_status") in ("MISSING", "URL_NOT_FOUND"))
    inconsistent_count = sum(1 for r in rows if r.get("last_validation_status") == "INCONSISTENT")
    expired_count = sum(1 for r in rows if r.get("is_expired"))
    ws.append([f"Conforme: {ok_count} | Nao Encontrado: {not_found_count} | Inconsistente: {inconsistent_count} | Vencidos: {expired_count}"])
    ws.append([])

    # Fetch stock data for all products
    stock_map: dict[str, dict] = {}
    try:
        with db() as (_conn2, _cur2):
            _cur2.execute("""
                SELECT sku, source, SUM(COALESCE(available, quantity, 0)) as qty
                FROM cert_stock GROUP BY sku, source
            """)
            for srow in _cur2.fetchall():
                sk = srow["sku"]
                if sk not in stock_map:
                    stock_map[sk] = {"cd": 0, "ecommerce": 0}
                if srow["source"] == "wms_biguacu":
                    stock_map[sk]["cd"] = srow["qty"] or 0
                else:
                    stock_map[sk]["ecommerce"] += srow["qty"] or 0
    except Exception as e:
        log.warning(f"Could not fetch stock data for export: {e}")

    # Headers
    headers = ["SKU", "Nome", "Marca", "Status", "Pontuacao", "Tipo Certificacao",
               "Texto Esperado", "Texto Encontrado", "URL", "Prazo Venda", "Vencido",
               "Estoque CD", "Estoque E-commerce", "Total Estoque"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data
    for r in rows:
        status_raw = r.get("last_validation_status") or ""
        is_exp = r.get("is_expired", False)
        if not status_raw and is_exp:
            status_raw = "EXPIRED"
        label = status_labels.get(status_raw, status_raw)
        score = r.get("last_validation_score")
        score_str = f"{score * 100:.0f}%" if score is not None else ""
        sku = r.get("sku", "")
        stock = stock_map.get(sku, {"cd": 0, "ecommerce": 0})
        row_data = [
            sku,
            r.get("name", ""),
            r.get("brand", ""),
            label,
            score_str,
            r.get("certification_type", ""),
            r.get("expected_cert_text", ""),
            r.get("actual_cert_text", ""),
            r.get("last_validation_url", ""),
            r.get("sale_deadline", ""),
            "Sim" if is_exp else "",
            stock["cd"],
            stock["ecommerce"],
            stock["cd"] + stock["ecommerce"],
        ]
        ws.append(row_data)
        row_idx = ws.max_row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
        # Color status
        status_cell = ws.cell(row=row_idx, column=4)
        if is_exp:
            status_cell.fill = expired_fill
        elif status_raw in status_fills:
            status_cell.fill = status_fills[status_raw]

    # Column widths
    col_widths = [15, 40, 18, 18, 12, 25, 40, 40, 50, 15, 10, 12, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A{header_row}:N{ws.max_row}"

    filename = f"produtos_certificacoes_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = REPORTS_DIR / filename
    wb.save(str(filepath))

    return FileResponse(
        str(filepath),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@app.post("/api/reports/export-stock")
@limiter.limit("5/minute")
def export_stock_report(request: Request, brand: str = Query("")):
    """Export detailed stock report with WMS locations as Excel."""
    if not DATABASE_URL:
        raise HTTPException(500, "Database not configured")

    now = datetime.now(timezone.utc)

    with db() as (conn, cur):
        conditions = []
        params: list = []
        if brand:
            conditions.append("cs.brand = %s")
            params.append(brand)

        where = "WHERE " + " AND ".join(conditions) if conditions else ""

        cur.execute(f"""
            SELECT cs.sku, cp.name, COALESCE(cp.brand, cs.brand) as brand,
                cs.source, cs.warehouse, cs.quantity, cs.available,
                cs.reserved, cs.in_transit, cs.situation, cs.storage_area,
                cp.last_validation_status, cp.sale_deadline, cp.is_expired,
                cs.synced_at
            FROM cert_stock cs
            LEFT JOIN cert_products cp ON cs.sku = cp.sku
            {where}
            ORDER BY cs.sku, cs.source, cs.warehouse
        """, params)
        rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque Detalhado"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Summary
    ws.append(["Relatório de Estoque Detalhado - CD Biguaçu + E-commerce"])
    ws.merge_cells("A1:L1")
    ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
    ws.append([f"Data: {now.strftime('%d/%m/%Y %H:%M')}"])
    ws.append([f"Total registros: {len(rows)}"])
    ws.append([])

    # Headers
    headers = ["SKU", "Nome", "Marca", "Origem", "Localização", "Quantidade",
               "Disponível", "Reserva", "Trânsito", "Situação", "Status Cert", "Prazo Venda"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Source labels
    source_labels = {
        "wms_biguacu": "CD Biguaçu (WMS)",
        "ecommerce_puket": "E-commerce Puket",
        "ecommerce_imaginarium": "E-commerce Imaginarium",
    }

    wms_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    ecom_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    for row_data in rows:
        source_raw = row_data.get("source", "")
        row_values = [
            row_data.get("sku", ""),
            row_data.get("name", ""),
            row_data.get("brand", ""),
            source_labels.get(source_raw, source_raw),
            (row_data.get("warehouse", "") or "").replace("CD ", ""),
            row_data.get("quantity", 0) or 0,
            row_data.get("available", 0) or 0,
            row_data.get("reserved", 0) or 0,
            row_data.get("in_transit", 0) or 0,
            row_data.get("situation", ""),
            row_data.get("last_validation_status", ""),
            row_data.get("sale_deadline", ""),
        ]
        ws.append(row_values)
        row_idx = ws.max_row
        fill = wms_fill if "wms" in source_raw else ecom_fill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.fill = fill

    # Column widths
    col_widths = [15, 45, 18, 25, 22, 12, 12, 10, 10, 20, 15, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A{header_row}:L{ws.max_row}"

    filename = f"estoque_detalhado_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = REPORTS_DIR / filename
    wb.save(str(filepath))

    return FileResponse(
        str(filepath),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@app.get("/api/reports")
def list_reports():
    reports = []
    if not REPORTS_DIR.exists():
        return reports
    for f in sorted(REPORTS_DIR.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True):
        if f.is_file():
            stat = f.stat()
            reports.append({
                "filename": f.name,
                "date": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
                "size_bytes": stat.st_size,
            })
    return reports


@app.get("/api/reports/{filename}/data")
def get_report_data(filename: str):
    if '..' in filename or '/' in filename or '\\' in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    filepath = REPORTS_DIR / filename
    if not filepath.exists() or not filepath.is_file():
        raise HTTPException(404, "Report not found")
    return json.loads(filepath.read_text())


@app.get("/api/reports/{filename}")
def download_report(filename: str, format: str = Query("xlsx")):
    if '..' in filename or '/' in filename or '\\' in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    filepath = REPORTS_DIR / filename
    if not filepath.exists() or not filepath.is_file():
        raise HTTPException(404, "Report not found")

    if format == "json":
        if filename.endswith(".xlsx"):
            raise HTTPException(400, "Este arquivo e binario (xlsx), nao pode ser baixado como JSON")
        return FileResponse(filepath, media_type="application/json", filename=filename)

    # If already an xlsx file, serve directly
    if filename.endswith(".xlsx"):
        return FileResponse(
            str(filepath),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )

    # Generate Excel from JSON report
    report_data = json.loads(filepath.read_text())
    products = report_data.get("products", report_data.get("results", []))
    summary = report_data.get("summary", {})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validação"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    status_fills = {
        "OK": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "URL_NOT_FOUND": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "INCONSISTENT": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "EXPIRED": PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid"),
    }
    status_labels = {
        "OK": "Conforme",
        "INCONSISTENT": "Inconsistente",
        "URL_NOT_FOUND": "Não Encontrado",
        "API_ERROR": "Erro de API",
        "NO_EXPECTED": "Sem Certificação",
        "EXPIRED": "Certificação Vencida",
    }

    # Summary row
    ws.append(["Relatório de Validação de Certificações"])
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=14, color="059669")
    ws.append([f"Data: {report_data.get('date', '')}"])
    ws.append([f"Total: {summary.get('total', len(products))} | OK: {summary.get('ok', 0)} | Ausente: {summary.get('missing', 0)} | Inconsistente: {summary.get('inconsistent', 0)} | Não Encontrado: {summary.get('not_found', 0)}"])
    ws.append([])

    # Fetch stock data for report
    stock_map: dict[str, dict] = {}
    try:
        with db() as (_conn3, _cur3):
            _cur3.execute("""
                SELECT sku, source, SUM(COALESCE(available, quantity, 0)) as qty
                FROM cert_stock GROUP BY sku, source
            """)
            for srow in _cur3.fetchall():
                sk = srow["sku"]
                if sk not in stock_map:
                    stock_map[sk] = {"cd": 0, "ecommerce": 0}
                if srow["source"] == "wms_biguacu":
                    stock_map[sk]["cd"] = srow["qty"] or 0
                else:
                    stock_map[sk]["ecommerce"] += srow["qty"] or 0
    except Exception as e:
        log.warning(f"Could not fetch stock data for report download: {e}")

    # Headers
    headers = ["SKU", "Nome", "Marca", "Status", "Pontuacao", "Texto Esperado", "Texto Encontrado", "URL", "Erro",
               "Estoque CD", "Estoque E-commerce", "Total Estoque"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for p in products:
        status_raw = p.get("status", "")
        status_label = status_labels.get(status_raw, status_raw)
        score = p.get("score")
        score_str = f"{score * 100:.0f}%" if score is not None else ""
        p_sku = p.get("sku", "")
        stock = stock_map.get(p_sku, {"cd": 0, "ecommerce": 0})
        row = [
            p_sku,
            p.get("name", ""),
            p.get("brand", ""),
            status_label,
            score_str,
            p.get("expected_cert_text", ""),
            p.get("actual_cert_text", ""),
            p.get("url", ""),
            p.get("error", ""),
            stock["cd"],
            stock["ecommerce"],
            stock["cd"] + stock["ecommerce"],
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
        # Color status cell
        status_cell = ws.cell(row=row_idx, column=4)
        if status_raw in status_fills:
            status_cell.fill = status_fills[status_raw]

    # Column widths
    col_widths = [15, 40, 18, 18, 12, 40, 40, 50, 40, 12, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Auto-filter
    ws.auto_filter.ref = f"A{header_row}:L{ws.max_row}"

    # Save to temp file
    excel_filename = filename.replace(".json", ".xlsx")
    excel_path = REPORTS_DIR / excel_filename
    wb.save(str(excel_path))

    return FileResponse(
        str(excel_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=excel_filename,
    )


# ---------------------------------------------------------------------------
# Licenciados (LI Tracking from Google Sheets)
# ---------------------------------------------------------------------------

def _ensure_li_tracking_table():
    """Create li_tracking table if it doesn't exist."""
    with db() as (conn, cur):
        # Create table without FK to import_processes to avoid startup ordering issues.
        # The process_id column is nullable and looked up at sync time.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS li_tracking (
                id SERIAL PRIMARY KEY,
                process_id INTEGER,
                process_code TEXT,
                ncm TEXT,
                orgao TEXT,
                supplier TEXT,
                item TEXT,
                description TEXT,
                status TEXT DEFAULT 'pending',
                lpco_number TEXT,
                valid_until DATE,
                created_at TIMESTAMPTZ DEFAULT NOW(),
                updated_at TIMESTAMPTZ DEFAULT NOW()
            )
        """)


def _read_licenciados_from_sheets() -> list[dict]:
    """Read 'Licenciados' worksheet from the shared spreadsheet."""
    client = _get_sheets_client()
    if not client:
        return []
    try:
        spreadsheet = client.open_by_key(SHEETS_SPREADSHEET_ID)
    except Exception as e:
        log.error(f"Failed to open spreadsheet for Licenciados: {e}")
        return []

    try:
        ws = spreadsheet.worksheet("Licenciados")
        rows = ws.get_all_values()
    except gspread.exceptions.WorksheetNotFound:
        log.info("Worksheet 'Licenciados' not found")
        return []
    except Exception as e:
        log.warning(f"Could not read worksheet 'Licenciados': {e}")
        return []

    if not rows or len(rows) < 2:
        return []

    headers = rows[0]
    ncm_col = _find_col_by_header(headers, "ncm")
    process_col = _find_col_by_header(headers, "processo", "process")
    orgao_col = _find_col_by_header(headers, "orgão", "orgao", "órgão")
    supplier_col = _find_col_by_header(headers, "fornecedor", "supplier")
    item_col = _find_col_by_header(headers, "item")
    desc_col = _find_col_by_header(headers, "descrição", "descricao", "description")
    status_col = _find_col_by_header(headers, "status")
    lpco_col = _find_col_by_header(headers, "lpco", "número lpco", "numero lpco")
    valid_col = _find_col_by_header(headers, "validade", "valid_until", "vencimento")

    items: list[dict] = []
    for row in rows[1:]:
        def get_val(col_idx: int | None) -> str:
            if col_idx is None or col_idx >= len(row):
                return ""
            return str(row[col_idx]).strip()

        process_code = get_val(process_col)
        if not process_code:
            continue

        valid_str = get_val(valid_col)
        valid_date = None
        if valid_str:
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
                try:
                    valid_date = datetime.strptime(valid_str, fmt).date()
                    break
                except ValueError:
                    continue

        items.append({
            "process_code": process_code,
            "ncm": get_val(ncm_col),
            "orgao": get_val(orgao_col),
            "supplier": get_val(supplier_col),
            "item": get_val(item_col),
            "description": get_val(desc_col),
            "status": get_val(status_col) or "pending",
            "lpco_number": get_val(lpco_col),
            "valid_until": valid_date,
        })

    log.info(f"Read {len(items)} licenciados from Google Sheets")
    return items


def sync_licenciados_to_db() -> dict:
    """Sync Licenciados from Google Sheets into li_tracking table."""
    items = _read_licenciados_from_sheets()
    if not items:
        return {"synced": 0, "error": "No licenciados found or Sheets not configured"}
    if not DATABASE_URL:
        return {"synced": 0, "error": "Database not configured"}

    _ensure_li_tracking_table()

    try:
        with db() as (conn, cur):
            # Build process_code -> process_id lookup
            cur.execute("SELECT id, process_code FROM import_processes")
            process_map = {r["process_code"]: r["id"] for r in cur.fetchall()}

            synced = 0
            for item in items:
                process_id = process_map.get(item["process_code"])

                cur.execute(
                    """
                    INSERT INTO li_tracking (process_id, process_code, ncm, orgao, supplier, item,
                                             description, status, lpco_number, valid_until, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                    ON CONFLICT ON CONSTRAINT li_tracking_process_code_ncm_item_key
                    DO UPDATE SET
                        process_id = COALESCE(EXCLUDED.process_id, li_tracking.process_id),
                        orgao = COALESCE(NULLIF(EXCLUDED.orgao, ''), li_tracking.orgao),
                        supplier = COALESCE(NULLIF(EXCLUDED.supplier, ''), li_tracking.supplier),
                        description = COALESCE(NULLIF(EXCLUDED.description, ''), li_tracking.description),
                        status = COALESCE(NULLIF(EXCLUDED.status, ''), li_tracking.status),
                        lpco_number = COALESCE(NULLIF(EXCLUDED.lpco_number, ''), li_tracking.lpco_number),
                        valid_until = COALESCE(EXCLUDED.valid_until, li_tracking.valid_until),
                        updated_at = NOW()
                    """,
                    [process_id, item["process_code"], item["ncm"], item["orgao"],
                     item["supplier"], item["item"], item["description"],
                     item["status"], item["lpco_number"], item["valid_until"]],
                )
                synced += 1

        return {"synced": synced, "total_rows": len(items)}
    except Exception as e:
        log.error(f"Failed to sync licenciados to DB: {e}")
        # If unique constraint doesn't exist, create it and retry (once only)
        if "li_tracking_process_code_ncm_item_key" in str(e):
            try:
                with db() as (conn, cur):
                    cur.execute("""
                        ALTER TABLE li_tracking
                        ADD CONSTRAINT li_tracking_process_code_ncm_item_key
                        UNIQUE (process_code, ncm, item)
                    """)
                # Retry the sync inline (no recursion)
                with db() as (conn, cur):
                    cur.execute("SELECT id, process_code FROM import_processes")
                    process_map = {r["process_code"]: r["id"] for r in cur.fetchall()}
                    synced = 0
                    for item in items:
                        process_id = process_map.get(item["process_code"])
                        cur.execute(
                            """
                            INSERT INTO li_tracking (process_id, process_code, ncm, orgao, supplier, item,
                                                     description, status, lpco_number, valid_until, updated_at)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                            ON CONFLICT ON CONSTRAINT li_tracking_process_code_ncm_item_key
                            DO UPDATE SET
                                process_id = COALESCE(EXCLUDED.process_id, li_tracking.process_id),
                                orgao = COALESCE(NULLIF(EXCLUDED.orgao, ''), li_tracking.orgao),
                                supplier = COALESCE(NULLIF(EXCLUDED.supplier, ''), li_tracking.supplier),
                                description = COALESCE(NULLIF(EXCLUDED.description, ''), li_tracking.description),
                                status = COALESCE(NULLIF(EXCLUDED.status, ''), li_tracking.status),
                                lpco_number = COALESCE(NULLIF(EXCLUDED.lpco_number, ''), li_tracking.lpco_number),
                                valid_until = COALESCE(EXCLUDED.valid_until, li_tracking.valid_until),
                                updated_at = NOW()
                            """,
                            [process_id, item["process_code"], item["ncm"], item["orgao"],
                             item["supplier"], item["item"], item["description"],
                             item["status"], item["lpco_number"], item["valid_until"]],
                        )
                        synced += 1
                return {"synced": synced, "total_rows": len(items), "note": "constraint created and retried"}
            except Exception as retry_err:
                log.error(f"Retry after constraint creation failed: {retry_err}")
        return {"synced": 0, "error": str(e)}


@app.post("/api/sync-licenciados")
@limiter.limit("5/minute")
def trigger_sync_licenciados(request: Request):
    if not SHEETS_CLIENT_EMAIL or not SHEETS_PRIVATE_KEY:
        raise HTTPException(400, "Google Sheets credentials not configured")
    result = sync_licenciados_to_db()
    if result.get("error"):
        raise HTTPException(500, result["error"])
    return result


@app.get("/api/licenciados")
def list_licenciados(
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=100),
    search: str = Query(""),
    status: str = Query(""),
    start_date: str = Query(""),
    end_date: str = Query(""),
):
    if not DATABASE_URL:
        return {"items": [], "total": 0, "page": 1, "per_page": per_page, "total_pages": 0}

    _ensure_li_tracking_table()

    with db() as (conn, cur):
        conditions: list[str] = []
        params: list = []

        if search:
            conditions.append("(process_code ILIKE %s OR ncm ILIKE %s OR description ILIKE %s)")
            params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
        if status:
            conditions.append("status = %s")
            params.append(status)
        if start_date:
            conditions.append("created_at >= %s::date")
            params.append(start_date)
        if end_date:
            conditions.append("created_at < (%s::date + interval '1 day')")
            params.append(end_date)

        where = "WHERE " + " AND ".join(conditions) if conditions else ""

        cur.execute(f"SELECT COUNT(*) as cnt FROM li_tracking {where}", params)
        total = cur.fetchone()["cnt"]

        offset = (page - 1) * per_page
        cur.execute(
            f"""SELECT * FROM li_tracking {where}
                ORDER BY updated_at DESC NULLS LAST
                LIMIT %s OFFSET %s""",
            params + [per_page, offset],
        )
        rows = []
        for r in cur.fetchall():
            row = dict(r)
            for dtfield in ("created_at", "updated_at", "valid_until"):
                if row.get(dtfield):
                    row[dtfield] = row[dtfield].isoformat() if hasattr(row[dtfield], 'isoformat') else str(row[dtfield])
            rows.append(row)

        total_pages = max(1, (total + per_page - 1) // per_page)

        return {
            "items": rows,
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": total_pages,
        }


# ---------------------------------------------------------------------------
# Stock / Inventory Integration
# ---------------------------------------------------------------------------

def _fetch_wms_stock():
    """Fetch stock from WMS Biguacu (Oracle) using thick mode for older Oracle servers."""
    import oracledb
    try:
        oracledb.init_oracle_client(lib_dir="/opt/oracle/instantclient_23_7")
    except Exception:
        pass  # Already initialized

    host = os.environ.get("WMS_ORACLE_HOST", "192.168.168.10")
    port = int(os.environ.get("WMS_ORACLE_PORT", "1521"))
    sid = os.environ.get("WMS_ORACLE_SID", "WMS")
    user = os.environ.get("WMS_ORACLE_USER", "wisweb")
    password = os.environ.get("WMS_ORACLE_PASS", "wisweb")

    dsn = oracledb.makedsn(host, port, sid=sid)
    with oracledb.connect(user=user, password=password, dsn=dsn) as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT te.CD_PRODUTO, tcp.DS_PRODUTO,
                tta.DS_AREA_ARMAZ AS Area,
                ts.DS_SITUACAO AS Situacao,
                SUM(te.QT_ESTOQUE) AS Estoque,
                SUM(te.QT_RESERVA_SAIDA) AS Reserva,
                SUM(te.QT_TRANSITO_ENTRADA) AS Transito,
                SUM(te.QT_TRANSITO_ENTRADA + te.QT_ESTOQUE - te.QT_RESERVA_SAIDA) AS Disponivel
            FROM T_ESTOQUE te
            JOIN T_CAB_PRODUTO tcp ON te.CD_PRODUTO = tcp.CD_PRODUTO
            JOIN T_ENDERECO_ESTOQUE tee ON te.CD_ENDERECO = tee.CD_ENDERECO
            JOIN T_SITUACAO ts ON ts.CD_SITUACAO = tee.CD_SITUACAO
            JOIN T_TIPO_AREA tta ON tta.CD_AREA_ARMAZ = tee.CD_AREA_ARMAZ
            GROUP BY te.CD_PRODUTO, tcp.DS_PRODUTO, tta.DS_AREA_ARMAZ, ts.DS_SITUACAO
        """)
        columns = [col[0] for col in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _fetch_ecommerce_stock(brand: str):
    """Fetch e-commerce stock from SQL Server ERP (Extrema MG)."""
    import pymssql

    if brand.lower() in ('puket', 'puket_escolares'):
        host = os.environ.get("ERP_PUKET_HOST", "db01.grupounico.com")
        db_name = os.environ.get("ERP_PUKET_DB", "DB_puket")
        filial = "EXTREMA - MG"
        where = f"filial = '{filial}'"
    else:
        host = os.environ.get("ERP_IMG_HOST", "db02.grupounico.com")
        db_name = os.environ.get("ERP_IMG_DB", "Grupo_Imaginarium")
        where = "filial LIKE '%IMAGINARIUM EXTREMA MG%'"

    user = os.environ.get("ERP_MSSQL_USER", "nicolas.matsuda")
    password = os.environ.get("ERP_MSSQL_PASS", "")

    with pymssql.connect(host, user, password, db_name, timeout=15, login_timeout=10) as conn:
        cursor = conn.cursor(as_dict=True)
        cursor.execute(f"SELECT * FROM estoque_produtos WHERE {where}")
        return cursor.fetchall()


@app.post("/api/sync-stock")
@limiter.limit("5/minute")
def sync_stock(request: Request):
    """Sync stock data from WMS + ERP into cert_stock table."""
    if not DATABASE_URL:
        raise HTTPException(500, "Database not configured")

    results = {"wms": 0, "ecommerce_puket": 0, "ecommerce_imaginarium": 0, "errors": []}

    now = datetime.now(timezone.utc).isoformat()

    # 1. WMS Oracle (by storage area)
    try:
        wms_rows = _fetch_wms_stock()
        with db() as (conn, cur):
            # Clear old WMS data before fresh insert
            cur.execute("DELETE FROM cert_stock WHERE source = 'wms_biguacu'")
            for row in wms_rows:
                sku = str(row.get("CD_PRODUTO", "")).strip()
                if not sku:
                    continue
                area = str(row.get("AREA", "")).strip() or "Geral"
                situation = str(row.get("SITUACAO", "")).strip() or ""
                warehouse_name = f"CD {area}"
                cur.execute("""
                    INSERT INTO cert_stock (sku, source, warehouse, quantity, available, reserved, in_transit, situation, storage_area, synced_at)
                    VALUES (%s, 'wms_biguacu', %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (sku, source, warehouse)
                    DO UPDATE SET quantity=EXCLUDED.quantity, available=EXCLUDED.available,
                        reserved=EXCLUDED.reserved, in_transit=EXCLUDED.in_transit,
                        situation=EXCLUDED.situation, storage_area=EXCLUDED.storage_area, synced_at=EXCLUDED.synced_at
                """, (sku, warehouse_name, int(row.get("ESTOQUE", 0) or 0), int(row.get("DISPONIVEL", 0) or 0),
                      int(row.get("RESERVA", 0) or 0), int(row.get("TRANSITO", 0) or 0),
                      situation, area, now))
        results["wms"] = len(wms_rows)
    except Exception as e:
        results["errors"].append(f"WMS Oracle: {str(e)}")

    # 2. ERP Puket
    try:
        puket_rows = _fetch_ecommerce_stock("puket")
        with db() as (conn, cur):
            for row in puket_rows:
                sku = str(row.get("PRODUTO", row.get("produto", ""))).strip()
                qty = int(row.get("ESTOQUE", row.get("estoque", 0)) or 0)
                if not sku:
                    continue
                cur.execute("""
                    INSERT INTO cert_stock (sku, brand, source, warehouse, quantity, available, synced_at)
                    VALUES (%s, 'puket', 'ecommerce_puket', 'Extrema MG', %s, %s, %s)
                    ON CONFLICT (sku, source, warehouse)
                    DO UPDATE SET quantity=EXCLUDED.quantity, available=EXCLUDED.available, synced_at=EXCLUDED.synced_at
                """, (sku, qty, qty, now))
        results["ecommerce_puket"] = len(puket_rows)
    except Exception as e:
        results["errors"].append(f"ERP Puket: {str(e)}")

    # 3. ERP Imaginarium
    try:
        img_rows = _fetch_ecommerce_stock("imaginarium")
        with db() as (conn, cur):
            for row in img_rows:
                sku = str(row.get("PRODUTO", row.get("produto", ""))).strip()
                qty = int(row.get("ESTOQUE", row.get("estoque", 0)) or 0)
                if not sku:
                    continue
                cur.execute("""
                    INSERT INTO cert_stock (sku, brand, source, warehouse, quantity, available, synced_at)
                    VALUES (%s, 'imaginarium', 'ecommerce_imaginarium', 'Extrema MG', %s, %s, %s)
                    ON CONFLICT (sku, source, warehouse)
                    DO UPDATE SET quantity=EXCLUDED.quantity, available=EXCLUDED.available, synced_at=EXCLUDED.synced_at
                """, (sku, qty, qty, now))
        results["ecommerce_imaginarium"] = len(img_rows)
    except Exception as e:
        results["errors"].append(f"ERP Imaginarium: {str(e)}")

    return results


@app.get("/api/stock/{sku}")
def get_stock_by_sku(sku: str):
    """Get stock data for a specific SKU across all sources."""
    if not DATABASE_URL:
        raise HTTPException(500, "Database not configured")

    with db() as (conn, cur):
        cur.execute(
            "SELECT source, warehouse, quantity, available, reserved, in_transit, synced_at FROM cert_stock WHERE sku = %s",
            (sku,),
        )
        rows = [dict(r) for r in cur.fetchall()]

    result = {"sku": sku, "stock": [], "total_quantity": 0, "total_available": 0}
    for item in rows:
        # Convert datetime to string for JSON serialization
        if item.get("synced_at") and hasattr(item["synced_at"], "isoformat"):
            item["synced_at"] = item["synced_at"].isoformat()
        result["stock"].append(item)
        result["total_quantity"] += item.get("quantity", 0) or 0
        result["total_available"] += item.get("available", 0) or 0
    return result
